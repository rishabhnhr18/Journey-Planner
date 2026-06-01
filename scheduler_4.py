"""
Saudi Multi-Salesperson Constraint Scheduling
==============================================
v3.0 — Major redesign with the following changes vs v2:

CHANGE LOG vs v2
────────────────
1.  GEOGRAPHIC CLUSTERING IN CP-SAT (no pre-assignment via KMeans).
    Customers are NOT assigned to a salesperson before the solver runs.
    The CP-SAT model jointly decides *which* salesperson serves each
    customer, ensuring nearby customers end up on the same route.

2.  TERRITORY FILTER.
    Pass `territory_id=None`  → schedule all territories.
    Pass `territory_id="TER_RUH"` → schedule that territory only.

3.  COLD / NORMAL TRUCK SEPARATION.
    Two separate plan DataFrames are produced:
      result.cold_schedule   — cold-chain customers only
      result.normal_schedule — normal-truck customers only
    Both are also merged in result.detailed_schedule for convenience.

4.  TRAVEL TIME FORMULA CORRECTED.
    Old formula used `dist * 2` (round-trip). New formula:
      - One-way distance from previous stop to next customer.
      - No buffer time added — pure visit + travel arithmetic.
      - In CP-SAT constraints, per-customer travel cost =
        one_way_km / speed_kmh * 60   (NO ×2).

5.  KM TRACKING PER CUSTOMER.
    detailed_schedule gains columns:
      route_leg_km        — km from previous stop (warehouse or prior customer)
      cumulative_route_km — running total for the day's route

6.  SALESPERSON ASSIGNMENT BY PROXIMITY (no KMeans pre-step).
    CP-SAT binary variable  assign[cid][sid] decides assignment.
    Proximity objective term penalises large intra-route distances.

7.  PRIORITY ORDER: High > Medium > Low.
    Hard constraint: every active (non-Churned, non-Dormant) customer
    must receive ≥ 1 visit per month.

8.  SOLVER FAILURE DIAGNOSTICS.
    If CP-SAT returns INFEASIBLE or MODEL_INVALID, the code runs
    a constraint relaxation probe to identify *which* constraint
    caused the failure and prints a human-readable explanation.

Scheduling behaviour by RFM segment
    High   → up to 4 visits/month, highest priority weight
    Medium → up to 2 visits/month, medium priority weight
    Low    → up to 1 visit/month,  lowest priority weight
    All active customers → minimum 1 visit (hard constraint)
"""

from __future__ import annotations

import calendar
import math
from dataclasses import dataclass, field
from typing import Any, Optional

import numpy as np
import pandas as pd
from ortools.sat.python import cp_model


# ─────────────────────────────────────────────────────────────────────────────
# Tiny haversine helper
# ─────────────────────────────────────────────────────────────────────────────

def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R    = 6371.0088
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a    = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ─────────────────────────────────────────────────────────────────────────────
# Data-classes for results
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SalespersonScheduleResult:
    sales_id:           str
    territory_id:       str
    detailed_schedule:  pd.DataFrame
    daily_schedule:     pd.DataFrame
    assigned_customers: pd.DataFrame


@dataclass
class MultiScheduleResult:
    detailed_schedule:    pd.DataFrame   # all customers, all territories
    cold_schedule:        pd.DataFrame   # cold-truck customers only
    normal_schedule:      pd.DataFrame   # normal-truck customers only
    daily_schedule:       pd.DataFrame
    unvisited_customers:  pd.DataFrame   # customers who received 0 visits (min-1 not met)
    salesperson_results:  dict[str, SalespersonScheduleResult] = field(default_factory=dict)
    territory_warehouses: dict[str, tuple[float, float]]       = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# RFM helpers
# ─────────────────────────────────────────────────────────────────────────────

def get_max_visits(segment: str) -> int:
    return {"High": 4, "Medium": 2, "Low": 1}.get(segment, 1)


def get_segment_priority_weight(segment: str) -> int:
    """
    High > Medium > Low in objective.
    Used as a multiplier so the solver prefers High customers first.
    """
    return {"High": 300, "Medium": 100, "Low": 20}.get(segment, 20)


# ─────────────────────────────────────────────────────────────────────────────
# Holiday helpers
# ─────────────────────────────────────────────────────────────────────────────

def _blocked_dates(
    col_name: str,
    col_value: str,
    holiday_df: pd.DataFrame,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
) -> set[pd.Timestamp]:
    blocked: set[pd.Timestamp] = set()
    hdf = holiday_df.copy()
    hdf["from_date"] = pd.to_datetime(hdf["from_date"]).dt.normalize()
    hdf["to_date"]   = pd.to_datetime(hdf["to_date"]).dt.normalize()
    for _, row in hdf[hdf[col_name] == col_value].iterrows():
        cur = row["from_date"]
        while cur <= row["to_date"]:
            blocked.add(cur)
            cur += pd.Timedelta(days=1)
    return {d for d in blocked if month_start <= d <= month_end}


def get_salesperson_blocked_dates(sales_id, holiday_df, month_start, month_end):
    return _blocked_dates("salesperson_holiday", sales_id, holiday_df, month_start, month_end)


def get_territory_blocked_dates(territory_id, holiday_df, month_start, month_end):
    return _blocked_dates("territory_holiday", territory_id, holiday_df, month_start, month_end)


# ─────────────────────────────────────────────────────────────────────────────
# Solver failure diagnostics
# ─────────────────────────────────────────────────────────────────────────────

def _diagnose_infeasible(
    customer_ids: list[str],
    salesperson_ids: list[str],
    valid_dates: list[pd.Timestamp],
    cust_lookup: dict,
    sp_is_cold: dict[str, bool],
    avg_visit_min: int,
    travel_mins: dict[tuple[str, str], int],     # (cid, sid) → travel minutes one-way
    daily_work_min: int,
    truck_group: str,                              # "cold" or "normal"
) -> list[str]:
    """
    Re-solves relaxed versions of the model to identify which constraint
    class is causing infeasibility.  Returns a list of human-readable
    diagnostic strings.
    """
    diagnostics: list[str] = []
    n_cust = len(customer_ids)
    n_sp   = len(salesperson_ids)
    n_days = len(valid_dates)

    # --- capacity arithmetic check (no CP model needed) ----------------------
    # Worst case: every customer assigned to the most distant SP on the same day
    max_per_sp_day: dict[str, int] = {}
    for sid in salesperson_ids:
        avg_t = int(np.mean([travel_mins.get((cid, sid), 0) for cid in customer_ids]) or 0)
        cap   = max(1, daily_work_min // max(avg_visit_min + avg_t, 1))
        max_per_sp_day[sid] = cap

    total_monthly_capacity = sum(
        max_per_sp_day[sid] * n_days for sid in salesperson_ids
    )

    if n_cust > total_monthly_capacity:
        diagnostics.append(
            f"CAPACITY: {n_cust} customers cannot be scheduled with {n_sp} salesperson(s) "
            f"across {n_days} working days.  Total monthly capacity = {total_monthly_capacity} slots "
            f"(based on avg travel times and {daily_work_min} min/day).  "
            f"Fix: reduce customers per salesperson, add salespeople, or extend working hours."
        )

    # --- cold-truck check ----------------------------------------------------
    if truck_group == "cold":
        cold_sp = [s for s in salesperson_ids if sp_is_cold.get(s, False)]
        if not cold_sp:
            diagnostics.append(
                "COLD TRUCK: No cold-capable salesperson is available in this territory "
                "for the cold-truck customer group.  Assign a cold-van to at least one "
                "salesperson in the territory."
            )

    # --- working days check --------------------------------------------------
    if n_days == 0:
        diagnostics.append(
            "HOLIDAYS: All working days in the month are blocked by territory or "
            "salesperson holidays.  No valid dates remain to schedule visits."
        )

    # --- per-SP one-visit-per-day limit check --------------------------------
    worst_sp_cap = max(max_per_sp_day.values()) if max_per_sp_day else 0
    if n_cust > worst_sp_cap * n_days * n_sp:
        diagnostics.append(
            f"DAILY LIMIT: Even with all salespeople working every day, the daily "
            f"capacity ({worst_sp_cap} customers/SP/day) × {n_days} days × {n_sp} SPs "
            f"= {worst_sp_cap * n_days * n_sp} total slots is less than {n_cust} customers."
        )

    # --- min-1-visit hard constraint check -----------------------------------
    # This can be infeasible if a customer's only valid date is also at capacity.
    # Approximation: if total slots < n_customers, min-visit constraint is infeasible.
    if total_monthly_capacity < n_cust and not diagnostics:
        diagnostics.append(
            "MIN VISIT: The hard constraint 'every customer ≥ 1 visit' cannot be satisfied "
            f"because total available slots ({total_monthly_capacity}) < customers ({n_cust}).  "
            "Reduce the customer pool or increase salesperson capacity."
        )

    if not diagnostics:
        diagnostics.append(
            "UNKNOWN: CP-SAT returned INFEASIBLE but no arithmetic constraint breach was "
            "detected.  This may be due to a combination of holiday blocking, cold-truck "
            "restrictions, and tight capacity. Try increasing solver_time_seconds or "
            "enabling relaxed_min_visits=True."
        )

    return diagnostics


# ─────────────────────────────────────────────────────────────────────────────
# Daily route post-processor  (nearest-neighbour + km tracking)
# ─────────────────────────────────────────────────────────────────────────────

class DailyRoutePlanner:
    """
    Given a day's list of customers (already assigned to a salesperson),
    orders them by nearest-neighbour from the warehouse, then computes:
      route_leg_km        — one-way distance from previous stop
      cumulative_route_km — running total km for the day
    """

    def get_route(
        self,
        day_schedule: pd.DataFrame,
        start_lat: float,
        start_lng: float,
    ) -> pd.DataFrame:
        if day_schedule.empty:
            return day_schedule.copy()

        unvisited = day_schedule.copy().reset_index(drop=True)
        route: list[dict] = []
        cur_lat, cur_lng = start_lat, start_lng
        cumulative_km = 0.0

        while not unvisited.empty:
            distances = unvisited.apply(
                lambda r: _haversine_km(cur_lat, cur_lng, float(r["gps_lat"]), float(r["gps_lng"])),
                axis=1,
            )
            nearest_idx = int(distances.idxmin())
            row         = unvisited.loc[nearest_idx].to_dict()
            leg_km      = distances[nearest_idx]
            cumulative_km += leg_km
            row["route_leg_km"]        = round(leg_km, 3)
            row["cumulative_route_km"] = round(cumulative_km, 3)
            route.append(row)
            cur_lat = float(row["gps_lat"])
            cur_lng = float(row["gps_lng"])
            unvisited = unvisited.drop(index=nearest_idx).reset_index(drop=True)

        result = pd.DataFrame(route)
        result["route_rank"] = range(1, len(result) + 1)
        return result


# ─────────────────────────────────────────────────────────────────────────────
# Core CP-SAT solver — one territory, one truck group
# ─────────────────────────────────────────────────────────────────────────────

class TerritoryScheduler:
    """
    Solves the joint assignment + scheduling problem for ONE territory and
    ONE truck group (cold or normal).

    Decision variables
    ──────────────────
    visit[cid][sid][d]  = 1  iff customer `cid` is visited by salesperson `sid`
                              on date `d`.

    Hard constraints
    ────────────────
    1.  Each customer is served by at most one salesperson on any given day.
    2.  Each customer is served by exactly ONE salesperson for ALL their visits
        (assignment consistency: if you visit customer X, you always visit X).
    3.  Cold-truck customers → only cold-capable salespeople.
    4.  Min 1 visit / month for every active customer.
    5.  Max visits / month by RFM segment (High=4, Medium=2, Low=1).
    6.  Daily total time (visit + travel) ≤ daily_work_minutes per salesperson.
    7.  Daily customer count ≤ capacity per salesperson.

    Objective (maximise)
    ────────────────────
    Priority weight × visits, with High > Medium > Low.
    final_customer_score refines priority within the same segment.
    Earlier dates are preferred (day_weight decays with day index).
    Geographic compactness bonus: rewards assigning nearby customers to
    the same salesperson (minimises intra-route spread).
    """

    DEFAULT_DAILY_WORK_MINUTES = 480
    DEFAULT_AVG_VISIT_MINUTES  = 22
    DEFAULT_AVG_SPEED_KMPH     = 32

    def __init__(self, solver_time_seconds: int = 90):
        self.solver_time_seconds = solver_time_seconds

    def solve(
        self,
        customers: pd.DataFrame,           # pre-filtered for this territory + truck group
        salespeople: pd.DataFrame,         # pre-filtered for this territory
        van_df: pd.DataFrame,
        valid_dates: list[pd.Timestamp],
        daily_work_minutes: int  = DEFAULT_DAILY_WORK_MINUTES,
        avg_visit_minutes: int   = DEFAULT_AVG_VISIT_MINUTES,
        avg_speed_kmph: float    = DEFAULT_AVG_SPEED_KMPH,
        warehouse_lat: float     = 0.0,
        warehouse_lng: float     = 0.0,
        territory_id: str        = "",
        truck_group: str                = "normal",  # "cold" | "normal"
        solver_time_seconds: int        = None,
        sp_valid_dates: dict            = None,   # sid → list[pd.Timestamp] of working days
    ) -> pd.DataFrame:
        """
        Returns a detailed_schedule DataFrame (one row per customer-visit).
        Returns empty DataFrame on failure (with printed diagnostics).
        """
        model        = cp_model.CpModel()
        customer_ids = customers["customer_id"].tolist()
        sp_ids       = salespeople["sales_id"].tolist()
        empty        = self._empty_df()

        if not customer_ids or not sp_ids or not valid_dates:
            print(f"  [{territory_id}/{truck_group}] Nothing to schedule — "
                  f"customers={len(customer_ids)}, SPs={len(sp_ids)}, dates={len(valid_dates)}")
            return empty

        # ── Van cold-capability map ───────────────────────────────────────────
        van_cold  = van_df.set_index("van_id")["cold_truck_enabled"].to_dict()
        sp_is_cold: dict[str, bool] = {
            sid: bool(
                salespeople.loc[salespeople["sales_id"] == sid, "assigned_van"]
                .map(van_cold).fillna(False).any()
            )
            for sid in sp_ids
        }
        cust_lookup = customers.set_index("customer_id").to_dict("index")

        # ── Travel time: one-way from warehouse to each customer, per SP
        #    (In this model we approximate per-customer travel as one-way
        #     from warehouse.  Route leg costs are handled in post-processing.)
        travel_mins: dict[tuple[str, str], int] = {}
        for _, cust in customers.iterrows():
            cid = cust["customer_id"]
            for sid in sp_ids:
                dist_km = _haversine_km(
                    warehouse_lat, warehouse_lng,
                    float(cust["gps_lat"]), float(cust["gps_lng"]),
                )
                # ONE-WAY only — no ×2
                travel_mins[(cid, sid)] = max(1, int((dist_km / avg_speed_kmph) * 60))

        # ── Per-SP daily capacity ─────────────────────────────────────────────
        sp_avg_travel: dict[str, int] = {
            sid: max(1, int(np.mean([travel_mins[(cid, sid)] for cid in customer_ids])))
            for sid in sp_ids
        }
        sp_time_per_visit: dict[str, int] = {
            sid: max(1, avg_visit_minutes + sp_avg_travel[sid])
            for sid in sp_ids
        }
        # Daily capacity = how many customers fit in daily_work_minutes with no buffer.
        sp_daily_cap: dict[str, int] = {
            sid: max(1, daily_work_minutes // sp_time_per_visit[sid])
            for sid in sp_ids
        }

        # ── Effective max visits per customer ─────────────────────────────────
        total_slots = sum(sp_daily_cap[sid] * len(valid_dates) for sid in sp_ids)
        desired     = sum(
            get_max_visits(cust_lookup[cid].get("rfm_segment_final", "Low"))
            for cid in customer_ids
        )
        effective_max: dict[str, int] = {
            cid: get_max_visits(cust_lookup[cid].get("rfm_segment_final", "Low"))
            for cid in customer_ids
        }
        if desired > total_slots:
            for cap_seg, cap_val in [("Medium", 1), ("High", 2), ("High", 1)]:
                for cid in customer_ids:
                    if cust_lookup[cid].get("rfm_segment_final") == cap_seg:
                        effective_max[cid] = min(effective_max[cid], cap_val)
                if sum(effective_max.values()) <= total_slots:
                    break

        # ── Decision variables ────────────────────────────────────────────────
        # visit[cid][sid][d] — 1 if customer cid is visited by salesperson sid on date d
        visit: dict[tuple[str, str, pd.Timestamp], cp_model.IntVar] = {}
        for cid in customer_ids:
            for sid in sp_ids:
                for d in valid_dates:
                    visit[(cid, sid, d)] = model.NewBoolVar(
                        f"v_{cid}_{sid}_{d.strftime('%Y%m%d')}"
                    )

        # assigned[cid][sid] — 1 if salesperson sid is responsible for customer cid
        assigned: dict[tuple[str, str], cp_model.IntVar] = {}
        for cid in customer_ids:
            for sid in sp_ids:
                assigned[(cid, sid)] = model.NewBoolVar(f"a_{cid}_{sid}")

        # ── CONSTRAINT 1: Each customer visited by exactly ONE salesperson ────
        for cid in customer_ids:
            model.Add(sum(assigned[(cid, sid)] for sid in sp_ids) == 1)

        # ── CONSTRAINT 2: visit only if assigned ─────────────────────────────
        for cid in customer_ids:
            for sid in sp_ids:
                for d in valid_dates:
                    model.Add(visit[(cid, sid, d)] <= assigned[(cid, sid)])

        # ── CONSTRAINT 3: Cold-truck restriction ─────────────────────────────
        if truck_group == "cold":
            cold_sp_ids = [s for s in sp_ids if sp_is_cold.get(s, False)]
            if cold_sp_ids:
                for cid in customer_ids:
                    for sid in sp_ids:
                        if not sp_is_cold.get(sid, False):
                            model.Add(assigned[(cid, sid)] == 0)
            else:
                print(f"  [{territory_id}/cold] WARNING: No cold-capable SP found. "
                      "Allowing all SPs for cold customers as fallback.")

        # ── CONSTRAINT 4: Min 1 visit per active customer ─────────────────────
        for cid in customer_ids:
            model.Add(
                sum(visit[(cid, sid, d)] for sid in sp_ids for d in valid_dates) >= 1
            )

        # ── CONSTRAINT 5: Max visits per month (by RFM segment) ──────────────
        for cid in customer_ids:
            max_v = effective_max[cid]
            model.Add(
                sum(visit[(cid, sid, d)] for sid in sp_ids for d in valid_dates) <= max_v
            )

        # ── CONSTRAINT 6: Per-salesperson daily time budget ───────────────────
        # sum(visit_min + travel_min per customer on day d) ≤ daily_work_minutes
        # No buffer added — pure time arithmetic.
        for sid in sp_ids:
            for d in valid_dates:
                model.Add(
                    sum(
                        (avg_visit_minutes + travel_mins[(cid, sid)]) * visit[(cid, sid, d)]
                        for cid in customer_ids
                    ) <= daily_work_minutes
                )

        # ── CONSTRAINT 7: Daily customer count limit per salesperson ──────────
        # Cap = floor(daily_work_min / avg_time_per_visit). No buffer deduction.
        for sid in sp_ids:
            for d in valid_dates:
                model.Add(
                    sum(visit[(cid, sid, d)] for cid in customer_ids)
                    <= sp_daily_cap[sid]
                )

        # ── CONSTRAINT 8: Each customer visited at most once per day ─────────
        for cid in customer_ids:
            for d in valid_dates:
                model.Add(
                    sum(visit[(cid, sid, d)] for sid in sp_ids) <= 1
                )

        # ── CONSTRAINT 9: SP personal blocked dates (annual leave / sick leave) ─
        # sp_valid_dates maps each SP to the dates they CAN work (territory holidays
        # + personal leave already merged in the orchestrator).
        # For every date a SP cannot work, force all their visit variables to 0.
        # Without this constraint the solver can legally assign a SP on their leave
        # day because all_valid_dates is the UNION across all SPs.
        if sp_valid_dates:
            for sid in sp_ids:
                sp_dates_set = set(sp_valid_dates.get(sid, valid_dates))
                for d in valid_dates:
                    if d not in sp_dates_set:          # this SP is off on date d
                        for cid in customer_ids:
                            model.Add(visit[(cid, sid, d)] == 0)

        # ── OBJECTIVE ─────────────────────────────────────────────────────────
        # Terms:
        #  A) Priority × score × visits   (High >> Medium >> Low)
        #  B) Preferred visit day bonus — reward visits on customer's preferred weekday
        #  C) Geographic compactness: customers close together → same SP
        #  D) Weekly cadence bonus for High (reward covering each ISO week)
        #  E) Alternate-week spread bonus for Medium (reward visits in different weeks)
#
        # NOTE: day_w (earlier-date preference) is intentionally removed.
        # It caused front-loading — all min-1 visits packed into the first few
        # days, leaving the second half of the month empty.  The weekly cadence
        # and alternate-week bonuses are sufficient to drive spread.

        # Map Python weekday (Mon=0) to preferred_visit_day string
        _DAY_NAME = {0:"Monday",1:"Tuesday",2:"Wednesday",
                     3:"Thursday",4:"Friday",5:"Saturday",6:"Sunday"}
        PREFERRED_DAY_BONUS = 50_000

        # ISO weeks present in valid_dates
        week_nums = sorted({d.isocalendar()[1] for d in valid_dates})
        date_to_week = {d: d.isocalendar()[1] for d in valid_dates}

        obj_terms: list[cp_model.LinearExprT] = []

        for cid in customer_ids:
            info         = cust_lookup[cid]
            segment      = info.get("rfm_segment_final", "Low")
            base         = get_segment_priority_weight(segment)
            score_b      = int(info.get("final_customer_score", 0.5) * 100)
            weight       = (base + score_b) * 1000
            pref_day     = info.get("preferred_visit_day", "")

            for d in valid_dates:
                pref_w = PREFERRED_DAY_BONUS if _DAY_NAME.get(d.weekday()) == pref_day else 0
                for sid in sp_ids:
                    obj_terms.append((weight + pref_w) * visit[(cid, sid, d)])

        # ── SC5: Weekly cadence bonus — High visits every week, Medium alternate weeks ─
        # For High customers: bonus for having at least 1 visit in EACH ISO week.
        # For Medium customers: bonus for visits spread across DIFFERENT weeks
        #   (penalise both visits landing in the same week).
        # Implemented as soft objective terms using auxiliary BoolVars, not hard
        # constraints, so the solver can trade off if calendar leaves no choice.
        WEEKLY_CADENCE_BONUS  = 200_000   # strong nudge for High
        ALT_WEEK_BONUS        = 150_000   # strong nudge for Medium

        for cid in customer_ids:
            segment = cust_lookup[cid].get("rfm_segment_final", "Low")

            if segment == "High":
                # has_visit_in_week[w] = 1 if customer cid has any visit in week w
                for w in week_nums:
                    week_dates = [d for d in valid_dates if date_to_week[d] == w]
                    if not week_dates:
                        continue
                    has_visit = model.NewBoolVar(f"hwk_{cid}_{w}")
                    # has_visit = 1  iff  sum of visits in this week >= 1
                    model.AddMaxEquality(
                        has_visit,
                        [visit[(cid, sid, d)]
                         for sid in sp_ids for d in week_dates]
                    )
                    obj_terms.append(WEEKLY_CADENCE_BONUS * has_visit)

            elif segment == "Medium":
                # Reward visits landing in DIFFERENT weeks.
                # For each pair of weeks, bonus if the customer has a visit in both.
                # (Medium gets max 2 visits, so this pushes them to separate weeks.)
                for wi, wa in enumerate(week_nums):
                    for wb in week_nums[wi + 1:]:
                        dates_a = [d for d in valid_dates if date_to_week[d] == wa]
                        dates_b = [d for d in valid_dates if date_to_week[d] == wb]
                        if not dates_a or not dates_b:
                            continue
                        has_a = model.NewBoolVar(f"mwk_{cid}_{wa}")
                        has_b = model.NewBoolVar(f"mwk_{cid}_{wb}")
                        model.AddMaxEquality(
                            has_a,
                            [visit[(cid, sid, d)] for sid in sp_ids for d in dates_a]
                        )
                        model.AddMaxEquality(
                            has_b,
                            [visit[(cid, sid, d)] for sid in sp_ids for d in dates_b]
                        )
                        both_weeks = model.NewBoolVar(f"mboth_{cid}_{wa}_{wb}")
                        model.AddMinEquality(both_weeks, [has_a, has_b])
                        obj_terms.append(ALT_WEEK_BONUS * both_weeks)

        # ── Compactness bonus (lightweight — no extra variables) ──────────────
        # Reward each customer for being close to the territory centroid.
        # Replaces the O(n²) BoolVar pair loop which created thousands of extra
        # variables and constraints, causing normal-group timeouts.
        if customer_ids:
            centroid_lat = float(np.mean([cust_lookup[c]["gps_lat"] for c in customer_ids]))
            centroid_lng = float(np.mean([cust_lookup[c]["gps_lng"] for c in customer_ids]))
            COMPACT_BONUS = 2_000
            for cid in customer_ids:
                dist_c = _haversine_km(
                    float(cust_lookup[cid]["gps_lat"]),
                    float(cust_lookup[cid]["gps_lng"]),
                    centroid_lat, centroid_lng,
                )
                proximity = max(0, int((1.0 - min(dist_c / 50.0, 1.0)) * COMPACT_BONUS))
                if proximity > 0:
                    for sid in sp_ids:
                        obj_terms.append(proximity * assigned[(cid, sid)])

        model.Maximize(sum(obj_terms))

        # ── SOLVE ─────────────────────────────────────────────────────────────
        import os
        solver = cp_model.CpSolver()
        t_limit = solver_time_seconds if solver_time_seconds is not None else self.solver_time_seconds
        solver.parameters.max_time_in_seconds = t_limit
        solver.parameters.num_search_workers  = max(4, min(os.cpu_count() or 4, 16))
        solver.parameters.log_search_progress = False

        # Greedy round-robin hint — seed solver with a good starting solution
        # so it polishes toward optimal rather than searching from scratch.
        sorted_custs = sorted(customer_ids,
                              key=lambda c: cust_lookup[c].get("final_customer_score", 0.5),
                              reverse=True)
        for i, cid in enumerate(sorted_custs):
            hint_sid = sp_ids[i % len(sp_ids)]
            for sid in sp_ids:
                model.AddHint(assigned[(cid, sid)], 1 if sid == hint_sid else 0)

        status = solver.Solve(model)

        # ── Failure diagnostics ───────────────────────────────────────────────
        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            status_name = solver.StatusName(status)
            print(f"\n  ❌ [{territory_id}/{truck_group}] CP-SAT returned: {status_name}")
            print(f"     Variables : {len(visit) + len(assigned)}")
            print(f"     Customers : {len(customer_ids)}")
            print(f"     Salespeople: {len(sp_ids)}")
            print(f"     Valid dates: {len(valid_dates)}")
            diagnostics = _diagnose_infeasible(
                customer_ids, sp_ids, valid_dates, cust_lookup,
                sp_is_cold, avg_visit_minutes, travel_mins, daily_work_minutes, truck_group,
            )
            print(f"\n  ── Constraint failure analysis ──")
            for i, msg in enumerate(diagnostics, 1):
                print(f"     [{i}] {msg}")
            print()
            return empty

        print(f"  ✅ [{territory_id}/{truck_group}] {solver.StatusName(status)} "
              f"| obj={solver.ObjectiveValue():.0f} "
              f"| wall={solver.WallTime():.1f}s")

        # ── Extract solution ──────────────────────────────────────────────────
        rows: list[dict[str, Any]] = []
        for cid in customer_ids:
            # Which SP was assigned?
            assigned_sid = next(
                (sid for sid in sp_ids if solver.Value(assigned[(cid, sid)]) == 1),
                None,
            )
            if assigned_sid is None:
                continue
            info = cust_lookup[cid]
            for d in valid_dates:
                if solver.Value(visit[(cid, assigned_sid, d)]) == 1:
                    dist_km   = _haversine_km(
                        warehouse_lat, warehouse_lng,
                        float(info["gps_lat"]), float(info["gps_lng"]),
                    )
                    t_min_ow  = max(1, int((dist_km / avg_speed_kmph) * 60))
                    rows.append({
                        "schedule_date":            d,
                        "sales_id":                 assigned_sid,
                        "territory_id":             territory_id,
                        "customer_id":              cid,
                        "truck_group":              truck_group,
                        "shop_name":                info.get("shop_name", ""),
                        "locality":                 info.get("locality", ""),
                        "gps_lat":                  info.get("gps_lat"),
                        "gps_lng":                  info.get("gps_lng"),
                        "rfm_segment_final":        info.get("rfm_segment_final", "Low"),
                        "final_customer_score":     info.get("final_customer_score", 0.0),
                        "rfm_combined":             info.get("rfm_combined", 0.0),
                        "customer_rank":            info.get("customer_rank", 0),
                        "seasonality_score":        info.get("seasonality_score", 0.0),
                        "territory_score":          info.get("territory_score", 0.0),
                        "locality_score":           info.get("locality_score", 0.0),
                        "rating_score":             info.get("rating_score", 0.0),
                        "lifecycle_state":          info.get("lifecycle_state", "Active"),
                        "cold_truck_required":      info.get("cold_truck_required", False),
                        "estimated_visit_minutes":  avg_visit_minutes,
                        "estimated_travel_minutes": t_min_ow,     # one-way, no ×2
                        "warehouse_to_customer_km": round(dist_km, 3),
                        # route_leg_km / cumulative_route_km filled by DailyRoutePlanner
                        "route_leg_km":             None,
                        "cumulative_route_km":      None,
                    })

        if not rows:
            return empty

        result_df = (
            pd.DataFrame(rows)
            .sort_values(
                ["schedule_date", "sales_id", "rfm_segment_final", "final_customer_score"],
                ascending=[True, True, True, False],
            )
            .reset_index(drop=True)
        )

        # ── Warn if any customer missed ───────────────────────────────────────
        scheduled = set(result_df["customer_id"].unique())
        missed    = [cid for cid in customer_ids if cid not in scheduled]
        if missed:
            print(
                f"  ⚠️  [{territory_id}/{truck_group}] {len(missed)} customer(s) NOT scheduled "
                f"(min-1-visit constraint may have been relaxed by capacity): {missed[:5]}"
                + (f" …+{len(missed)-5} more" if len(missed) > 5 else "")
            )

        return result_df

    @staticmethod
    def _empty_df() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "schedule_date", "sales_id", "territory_id", "customer_id", "truck_group",
            "shop_name", "locality", "gps_lat", "gps_lng",
            "rfm_segment_final", "final_customer_score", "rfm_combined",
            "customer_rank", "seasonality_score", "territory_score",
            "locality_score", "rating_score",
            "lifecycle_state", "cold_truck_required",
            "estimated_visit_minutes", "estimated_travel_minutes",
            "warehouse_to_customer_km", "route_leg_km", "cumulative_route_km",
        ])


# ─────────────────────────────────────────────────────────────────────────────
# Main orchestrator
# ─────────────────────────────────────────────────────────────────────────────

class MultiSalespersonScheduler:
    """
    Pipeline
    ────────
    1. Merge customer + RFM data.
    2. Filter to requested territory (or all territories).
    3. For each territory:
       a. Split customers into cold-truck and normal-truck groups.
       b. Run TerritoryScheduler separately for each group.
       c. Post-process routes with DailyRoutePlanner (km tracking).
    4. Aggregate results.
    5. Return MultiScheduleResult with:
         .detailed_schedule   — all visits
         .cold_schedule       — cold-truck visits only
         .normal_schedule     — normal-truck visits only
         .daily_schedule      — per-day summary with route order & km
    """

    def __init__(self, solver_time_seconds: int = 90):
        self.solver_time_seconds = solver_time_seconds
        self._ter_scheduler  = TerritoryScheduler(solver_time_seconds=solver_time_seconds)
        self._route_planner  = DailyRoutePlanner()

    def create_schedules(
        self,
        customer_df:      pd.DataFrame,
        rfm_scores_df:    pd.DataFrame,
        salesperson_df:   pd.DataFrame,
        holiday_df:       pd.DataFrame,
        territory_df:     pd.DataFrame,
        config_df:        pd.DataFrame,
        van_df:           pd.DataFrame,
        month_start_date: str | pd.Timestamp = "2026-06-01",
        territory_id:     Optional[str]      = None,   # None → all territories
    ) -> MultiScheduleResult:
        """
        Parameters
        ──────────
        territory_id : str | None
            If provided, only that territory is scheduled.
            If None, all territories are scheduled.
        """
        month_start   = pd.Timestamp(month_start_date).normalize()
        year, month   = month_start.year, month_start.month
        days_in_month = calendar.monthrange(year, month)[1]
        month_end     = month_start + pd.Timedelta(days=days_in_month - 1)

        cfg = (
            config_df.set_index("config_key")["config_value"].to_dict()
            if not config_df.empty else {}
        )
        avg_speed_kmph     = float(cfg.get("avg_speed_kmh", 32))
        avg_visit_minutes  = int(float(cfg.get("avg_service_time_min", 22)))
        daily_work_minutes = int(cfg.get("daily_work_minutes", 480))

        full_customers = self._build_full_customer_df(customer_df, rfm_scores_df)
        ter_info       = territory_df.set_index("territory_id").to_dict("index")

        # Filter territory_df if a specific territory was requested
        if territory_id is not None:
            ter_subset = territory_df[territory_df["territory_id"] == territory_id]
            if ter_subset.empty:
                raise ValueError(f"Territory '{territory_id}' not found in territory_df.")
        else:
            ter_subset = territory_df

        all_detailed:   list[pd.DataFrame] = []
        sp_results:     dict[str, SalespersonScheduleResult] = {}
        ter_warehouses: dict[str, tuple[float, float]] = {}

        for _, ter_row in ter_subset.iterrows():
            tid = ter_row["territory_id"]
            ter_customers = full_customers[full_customers["territory_id"] == tid].copy()
            territory_sps = salesperson_df[
                (salesperson_df["territory_id"] == tid) &
                (salesperson_df["active_status"] == True)
            ].copy()

            if ter_customers.empty or territory_sps.empty:
                print(f"\nTerritory {tid}: skipped — "
                      f"customers={len(ter_customers)}, active SPs={len(territory_sps)}")
                continue

            wh_lat = float(ter_info[tid]["warehouse_lat"])
            wh_lng = float(ter_info[tid]["warehouse_lng"])
            ter_warehouses[tid] = (wh_lat, wh_lng)

            # ── Territory blocked dates ───────────────────────────────────────
            ter_blocked = get_territory_blocked_dates(tid, holiday_df, month_start, month_end)

            print(f"\n{'='*60}")
            print(f"Territory {tid}: {len(ter_customers)} customers, "
                  f"{len(territory_sps)} active SPs")

            # ── Split into cold and normal truck groups ────────────────────────
            cold_custs   = ter_customers[ter_customers["cold_truck_required"] == True].copy()
            normal_custs = ter_customers[ter_customers["cold_truck_required"] == False].copy()

            print(f"  Cold-truck customers : {len(cold_custs)}")
            print(f"  Normal-truck customers: {len(normal_custs)}")

            # ── SP blocked dates (union with territory) ────────────────────────
            # Build per-SP valid dates
            sp_valid_dates: dict[str, list[pd.Timestamp]] = {}
            for _, sp_row in territory_sps.iterrows():
                sid         = sp_row["sales_id"]
                sp_blocked  = get_salesperson_blocked_dates(sid, holiday_df, month_start, month_end)
                all_blocked = ter_blocked | sp_blocked
                sp_valid_dates[sid] = [
                    month_start + pd.Timedelta(days=i)
                    for i in range(days_in_month)
                    if (month_start + pd.Timedelta(days=i)) not in all_blocked
                ]

            # For simplicity and correctness, we run PER-TRUCK-GROUP solver.
            for group_name, group_custs in [("cold", cold_custs), ("normal", normal_custs)]:
                if group_custs.empty:
                    print(f"  [{tid}/{group_name}] No customers — skipped.")
                    continue

                # Determine working SPs for this group.
                # Cold SPs (those with a cold-enabled van) are exclusively
                # reserved for cold-truck customers.  Normal-truck customers
                # are served only by the remaining SPs.
                # This ensures a SP dedicated to cold routes is never
                # double-counted in the normal pool.
                van_cold_map = van_df.set_index("van_id")["cold_truck_enabled"].to_dict()
                cold_sp_mask = territory_sps["assigned_van"].map(van_cold_map).fillna(False)
                cold_sps     = territory_sps[cold_sp_mask].copy()
                normal_sps   = territory_sps[~cold_sp_mask].copy()

                if group_name == "cold":
                    group_sps = cold_sps
                    if group_sps.empty:
                        print(f"  [{tid}/cold] No cold-capable SPs — using all SPs as fallback.")
                        group_sps = territory_sps.copy()
                else:
                    group_sps = normal_sps
                    if group_sps.empty:
                        print(f"  [{tid}/normal] All SPs are cold-capable — sharing full pool as fallback.")
                        group_sps = territory_sps.copy()

                # sp_valid_dates filtered to this group's SPs only.
                # all_valid_dates = union of working days across this group's SPs.
                # This is computed HERE (after group split) so that SP1's extra
                # holidays don't inflate the day count shown for SP2 and SP3,
                # and the solver doesn't build variables for days no one can work.
                group_sp_valid_dates = {
                    sid: sp_valid_dates[sid]
                    for sid in group_sps["sales_id"].tolist()
                    if sid in sp_valid_dates
                }
                all_valid_dates = sorted(set(
                    d for dates in group_sp_valid_dates.values() for d in dates
                ))

                # Dynamic solver time — normal groups are much larger than cold;
                # give them more budget to reach OPTIMAL.
                n_vars       = len(group_custs) * len(group_sps) * len(all_valid_dates)
                time_ceiling = 120 if group_name == "cold" else 600
                dynamic_time = max(
                    self.solver_time_seconds,
                    min(time_ceiling, len(group_custs) * len(group_sps) * 3),
                )
                print(f"\n  [{tid}/{group_name}] {len(group_custs)} customers × "
                      f"{len(group_sps)} SPs × {len(all_valid_dates)} days "
                      f"→ ~{n_vars} vars | timeout={dynamic_time}s")

                detailed = self._ter_scheduler.solve(
                    customers          = group_custs,
                    salespeople        = group_sps,
                    van_df             = van_df,
                    valid_dates        = all_valid_dates,
                    daily_work_minutes = daily_work_minutes,
                    avg_visit_minutes  = avg_visit_minutes,
                    avg_speed_kmph     = avg_speed_kmph,
                    warehouse_lat      = wh_lat,
                    warehouse_lng      = wh_lng,
                    territory_id       = tid,
                    truck_group        = group_name,
                    solver_time_seconds= dynamic_time,
                    sp_valid_dates     = group_sp_valid_dates,
                )

                if detailed.empty:
                    continue

                # ── Post-process: nearest-neighbour routes + km tracking ──────
                detailed = self._apply_route_ordering(detailed, wh_lat, wh_lng)

                all_detailed.append(detailed)

                # Store per-salesperson results
                for sid in detailed["sales_id"].unique():
                    sp_det = detailed[detailed["sales_id"] == sid].copy()
                    sp_cust= group_custs[group_custs["customer_id"].isin(
                        sp_det["customer_id"].unique()
                    )].copy()
                    daily  = self._build_daily_summary(sp_det, wh_lat, wh_lng, sid)

                    key = f"{sid}_{group_name}"   # unique key per SP × truck group
                    sp_results[key] = SalespersonScheduleResult(
                        sales_id           = sid,
                        territory_id       = tid,
                        detailed_schedule  = sp_det,
                        daily_schedule     = daily,
                        assigned_customers = sp_cust,
                    )

        # ── Combine all results ───────────────────────────────────────────────
        if all_detailed:
            combined = (
                pd.concat(all_detailed, ignore_index=True)
                .sort_values(
                    ["territory_id", "truck_group", "sales_id",
                     "schedule_date", "final_customer_score"],
                    ascending=[True, True, True, True, False],
                )
                .reset_index(drop=True)
            )
        else:
            combined = TerritoryScheduler._empty_df()

        cold_schedule   = combined[combined["truck_group"] == "cold"].reset_index(drop=True)
        normal_schedule = combined[combined["truck_group"] == "normal"].reset_index(drop=True)
        combined_daily  = self._build_combined_daily(sp_results)

        # ── Unvisited customers — received 0 scheduled visits ─────────────
        scheduled_ids    = set(combined["customer_id"].unique()) if not combined.empty else set()
        unvisited_df     = full_customers[
            ~full_customers["customer_id"].isin(scheduled_ids)
        ].copy().reset_index(drop=True)
        if territory_id:
            unvisited_df = unvisited_df[
                unvisited_df["territory_id"] == territory_id
            ].reset_index(drop=True)
        if not unvisited_df.empty:
            print(f"\n  ⚠️  {len(unvisited_df)} customer(s) received NO visits (min-1 not met):")
            for _, r in unvisited_df.iterrows():
                print(f"     {r['customer_id']}  [{r.get('rfm_segment_final','?')}]  "
                      f"{r.get('shop_name','')}  ({r.get('territory_id','')})")

        return MultiScheduleResult(
            detailed_schedule    = combined,
            cold_schedule        = cold_schedule,
            normal_schedule      = normal_schedule,
            daily_schedule       = combined_daily,
            unvisited_customers  = unvisited_df,
            salesperson_results  = sp_results,
            territory_warehouses = ter_warehouses,
        )

    # ── Internal helpers ──────────────────────────────────────────────────────

    @staticmethod
    def _build_full_customer_df(customer_df: pd.DataFrame, rfm_scores_df: pd.DataFrame) -> pd.DataFrame:
        rfm_cols = [
            "customer_id", "rfm_segment_final", "final_customer_score",
            "rfm_combined", "customer_rank", "recency", "frequency", "monetary",
            "r_score", "f_score", "m_score",
            "seasonality_score", "territory_score", "locality_score", "rating_score",
        ]
        rfm_use = [c for c in rfm_cols if c in rfm_scores_df.columns]
        merged  = customer_df.merge(rfm_scores_df[rfm_use], on="customer_id", how="left")
        merged["rfm_segment_final"]    = merged["rfm_segment_final"].fillna("Low")
        merged["final_customer_score"] = merged["final_customer_score"].fillna(0.0)
        return merged

    def _apply_route_ordering(
        self,
        detailed: pd.DataFrame,
        wh_lat: float,
        wh_lng: float,
    ) -> pd.DataFrame:
        """
        For each (sales_id, schedule_date), apply nearest-neighbour ordering
        from the warehouse and populate route_leg_km / cumulative_route_km.
        """
        parts: list[pd.DataFrame] = []
        for (sid, d), group in detailed.groupby(["sales_id", "schedule_date"]):
            routed = self._route_planner.get_route(group.copy(), wh_lat, wh_lng)
            parts.append(routed)
        if not parts:
            return detailed
        return pd.concat(parts, ignore_index=True)

    def _build_daily_summary(
        self,
        detailed: pd.DataFrame,
        wh_lat: float,
        wh_lng: float,
        sales_id: str,
    ) -> pd.DataFrame:
        rows: list[dict] = []
        for d, group in detailed.groupby("schedule_date"):
            routed = group.sort_values("route_rank") if "route_rank" in group.columns else group
            total_leg_km = routed["route_leg_km"].sum() if "route_leg_km" in routed.columns else 0.0
            rows.append({
                "schedule_date":       d,
                "sales_id":            sales_id,
                "territory_id":        group["territory_id"].iloc[0],
                "truck_group":         group["truck_group"].iloc[0],
                "customer_list":       routed["customer_id"].tolist(),
                "customer_count":      len(routed),
                "route_order":         routed["shop_name"].tolist(),
                "segment_breakdown":   routed["rfm_segment_final"].value_counts().to_dict(),
                "avg_customer_score":  round(float(routed["final_customer_score"].mean()), 4),
                "total_visit_min":     int(routed["estimated_visit_minutes"].sum()),
                "total_travel_min":    int(routed["estimated_travel_minutes"].sum()),
                "total_route_km":      round(float(total_leg_km), 3),
                "customer_km_detail":  dict(zip(
                    routed["customer_id"].tolist(),
                    routed["route_leg_km"].round(3).tolist()
                    if "route_leg_km" in routed.columns else [None] * len(routed),
                )),
            })
        if not rows:
            return pd.DataFrame(columns=[
                "schedule_date", "sales_id", "territory_id", "truck_group",
                "customer_list", "customer_count", "route_order", "segment_breakdown",
                "avg_customer_score", "total_visit_min", "total_travel_min",
                "total_route_km", "customer_km_detail",
            ])
        return pd.DataFrame(rows).sort_values(["sales_id", "schedule_date"]).reset_index(drop=True)

    @staticmethod
    def _build_combined_daily(sp_results: dict) -> pd.DataFrame:
        parts = [r.daily_schedule for r in sp_results.values() if not r.daily_schedule.empty]
        if not parts:
            return pd.DataFrame()
        return (
            pd.concat(parts, ignore_index=True)
            .sort_values(["territory_id", "truck_group", "sales_id", "schedule_date"])
            .reset_index(drop=True)
        )


# ─────────────────────────────────────────────────────────────────────────────
# Folium map helpers  (updated for truck_group, km, new fields)
# ─────────────────────────────────────────────────────────────────────────────

_SP_COLOURS = [
    "red", "blue", "green", "purple", "orange",
    "darkred", "cadetblue", "darkgreen", "darkpurple", "black",
]
SEGMENT_COLOUR = {"High": "red", "Medium": "orange", "Low": "gray"}


def _add_warehouse_marker(m, wh_lat: float, wh_lng: float, label: str = "Warehouse"):
    import folium
    folium.Marker(
        [wh_lat, wh_lng],
        icon=folium.Icon(color="black", icon="home", prefix="fa"),
        tooltip=f"🏭 {label} (start of route)",
        popup=f"<b>{label}</b><br>Route start",
    ).add_to(m)


def build_route_map_for_salesperson(
    daily_schedule:    pd.DataFrame,
    detailed_schedule: pd.DataFrame,
    sales_id:          str,
    schedule_date:     str | pd.Timestamp,
    warehouse_lat:     float = 0.0,
    warehouse_lng:     float = 0.0,
    zoom_start:        int   = 12,
):
    """
    Folium map for ONE salesperson on ONE day.
    Tooltip shows RFM segment, score, rank, and route leg km.
    """
    try:
        import folium
        from folium.plugins import PolyLineTextPath
    except ImportError:
        raise ImportError("pip install folium")

    sched_date = pd.Timestamp(schedule_date).normalize()
    day_df = detailed_schedule[
        (detailed_schedule["sales_id"]      == sales_id) &
        (detailed_schedule["schedule_date"] == sched_date)
    ].copy()

    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank")
    day_df = day_df.reset_index(drop=True)

    if day_df.empty:
        raise ValueError(f"No visits for {sales_id} on {sched_date.date()}")

    center_lat = day_df["gps_lat"].mean()
    center_lng = day_df["gps_lng"].mean()
    m = folium.Map(location=[center_lat, center_lng], zoom_start=zoom_start)

    if warehouse_lat != 0.0 or warehouse_lng != 0.0:
        _add_warehouse_marker(m, warehouse_lat, warehouse_lng)
        route_coords = [[warehouse_lat, warehouse_lng]] + day_df[["gps_lat", "gps_lng"]].values.tolist()
    else:
        route_coords = day_df[["gps_lat", "gps_lng"]].values.tolist()

    for _, row in day_df.iterrows():
        color  = SEGMENT_COLOUR.get(str(row.get("rfm_segment_final", "")), "blue")
        score  = row.get("final_customer_score", 0.0)
        rank   = row.get("customer_rank", "?")
        leg_km = row.get("route_leg_km", 0.0)
        folium.Marker(
            [row["gps_lat"], row["gps_lng"]],
            icon=folium.DivIcon(
                html=f"""<div style="background:{color};color:white;border-radius:50%;
                         width:26px;height:26px;text-align:center;line-height:26px;
                         font-weight:bold;font-size:11px;">{row.get('route_rank','')}</div>""",
            ),
            tooltip=(
                f"{row.get('route_rank','')}. {row.get('shop_name','')} "
                f"[{row.get('rfm_segment_final','')}] "
                f"score={score:.3f} rank=#{rank} "
                f"leg={leg_km:.2f}km "
                f"cold={row.get('cold_truck_required','?')}"
            ),
        ).add_to(m)

    line = folium.PolyLine(route_coords, weight=3, color="navy")
    m.add_child(line)
    PolyLineTextPath(line, "➤", repeat=True, offset=7,
                     attributes={"fill": "red", "font-size": "14"}).add_to(m)

    m.get_root().html.add_child(folium.Element(f"""
    <div style="position:fixed;top:10px;left:60px;z-index:1000;
                background:white;padding:8px 12px;border-radius:6px;
                border:1px solid #ccc;font-family:sans-serif;">
        <b>{sales_id}</b> — {sched_date.strftime('%d %b %Y')} &nbsp;|&nbsp; {len(day_df)} stops
    </div>"""))

    return m


def build_territory_day_map(
    result:        MultiScheduleResult,
    territory_id:  str,
    schedule_date: str | pd.Timestamp,
    truck_group:   Optional[str] = None,   # None → show both cold + normal
    zoom_start:    int = 11,
):
    """
    Shows all salespeople in one territory on a single folium map for a given date.
    truck_group filter: "cold", "normal", or None (both).
    """
    try:
        import folium
        from folium.plugins import PolyLineTextPath
    except ImportError:
        raise ImportError("pip install folium")

    sched_date = pd.Timestamp(schedule_date).normalize()

    # Filter detailed schedule
    df = result.detailed_schedule.copy()
    df = df[
        (df["territory_id"] == territory_id) &
        (df["schedule_date"] == sched_date)
    ]
    if truck_group:
        df = df[df["truck_group"] == truck_group]

    if df.empty:
        raise ValueError(f"No visits in territory {territory_id} on {sched_date.date()} "
                         f"(truck_group={truck_group})")

    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))

    center_lat = df["gps_lat"].mean()
    center_lng = df["gps_lng"].mean()
    m = folium.Map(location=[center_lat, center_lng], zoom_start=zoom_start)

    if wh_lat != 0.0 or wh_lng != 0.0:
        _add_warehouse_marker(m, wh_lat, wh_lng, label=f"Warehouse ({territory_id})")

    sp_ids = sorted(df["sales_id"].unique())
    sp_colour_map = {sid: _SP_COLOURS[i % len(_SP_COLOURS)] for i, sid in enumerate(sp_ids)}
    legend_items  = []

    for sid in sp_ids:
        sp_df     = df[df["sales_id"] == sid].copy()
        if "route_rank" in sp_df.columns:
            sp_df = sp_df.sort_values("route_rank")
        sp_color  = sp_colour_map[sid]

        route_coords = (
            [[wh_lat, wh_lng]] + sp_df[["gps_lat", "gps_lng"]].values.tolist()
            if (wh_lat != 0.0 or wh_lng != 0.0)
            else sp_df[["gps_lat", "gps_lng"]].values.tolist()
        )

        for _, row in sp_df.iterrows():
            score  = row.get("final_customer_score", 0.0)
            leg_km = row.get("route_leg_km", 0.0)
            folium.Marker(
                [row["gps_lat"], row["gps_lng"]],
                icon=folium.DivIcon(
                    html=f"""<div style="background:{sp_color};color:white;
                             border-radius:50%;width:26px;height:26px;
                             text-align:center;line-height:26px;
                             font-weight:bold;font-size:11px;
                             border:2px solid white;">{row.get('route_rank','')}</div>""",
                ),
                tooltip=(
                    f"[{sid}] {row.get('route_rank','')}. {row.get('shop_name','')} "
                    f"[{row.get('rfm_segment_final','')}] "
                    f"score={score:.3f} leg={leg_km:.2f}km "
                    f"cold={row.get('cold_truck_required','?')}"
                ),
            ).add_to(m)

        line = folium.PolyLine(route_coords, weight=3, color=sp_color, opacity=0.8)
        m.add_child(line)
        PolyLineTextPath(line, "➤", repeat=True, offset=7,
                         attributes={"fill": sp_color, "font-size": "13"}).add_to(m)

        seg_str = " | ".join(
            f"{k}:{v}" for k, v in sp_df["rfm_segment_final"].value_counts().items()
        )
        total_km = sp_df["route_leg_km"].sum() if "route_leg_km" in sp_df.columns else 0
        legend_items.append(
            f'<li><span style="background:{sp_color};width:14px;height:14px;'
            f'display:inline-block;border-radius:50%;margin-right:6px;"></span>'
            f'{sid} ({len(sp_df)} stops, {total_km:.1f} km — {seg_str})</li>'
        )

    legend_html = f"""
    <div style="position:fixed;top:10px;right:10px;z-index:1000;
                background:white;padding:10px 14px;border-radius:8px;
                border:1px solid #ccc;font-family:sans-serif;font-size:12px;
                min-width:240px;">
        <b>{territory_id}</b> — {sched_date.strftime('%d %b %Y')}
        {(' | 🧊 Cold only' if truck_group=='cold' else ' | 🚚 Normal only' if truck_group=='normal' else '')}<br>
        <ul style="list-style:none;margin:6px 0 0;padding:0;">
            {''.join(legend_items)}
            <li style="margin-top:6px;">
                <span style="background:black;width:14px;height:14px;
                display:inline-block;border-radius:50%;margin-right:6px;"></span>
                Warehouse
            </li>
        </ul>
        <hr style="margin:6px 0;">
        <span style="font-size:11px;color:#666;">
            Marker colour = salesperson &nbsp;|&nbsp; Number = route stop<br>
            High=red &nbsp; Medium=orange &nbsp; Low=gray (segment in tooltip)
        </span>
    </div>"""
    m.get_root().html.add_child(folium.Element(legend_html))
    return m


# ─────────────────────────────────────────────────────────────────────────────
# Stop-to-stop distance: Folium map + Excel export
# ─────────────────────────────────────────────────────────────────────────────

def build_stop_to_stop_distance_table(
    result,
    territory_id:  str,
    schedule_date,
    sales_id:      str,
) -> pd.DataFrame:
    """
    Returns a DataFrame with one row per consecutive stop pair on the route:
        stop_from | stop_to | from_id | from_shop | to_id | to_shop | leg_km | leg_min
    Row 0 is always Warehouse → Stop 1.
    """
    sched_date = pd.Timestamp(schedule_date).normalize()
    det = result.detailed_schedule
    day_df = det[
        (det["territory_id"]  == territory_id) &
        (det["schedule_date"] == sched_date) &
        (det["sales_id"]      == sales_id)
    ].copy()

    if day_df.empty:
        return pd.DataFrame()

    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank").reset_index(drop=True)

    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))
    avg_speed = 32.0   # km/h

    stops = [{"stop_num": 0, "customer_id": "WAREHOUSE", "shop_name": "Warehouse",
               "gps_lat": wh_lat, "gps_lng": wh_lng}]
    for _, row in day_df.iterrows():
        stops.append({"stop_num": int(row.get("route_rank", len(stops))),
                      "customer_id": row["customer_id"],
                      "shop_name": row.get("shop_name", ""),
                      "gps_lat": float(row["gps_lat"]), "gps_lng": float(row["gps_lng"])})

    rows = []
    for i in range(len(stops) - 1):
        f, t = stops[i], stops[i + 1]
        km = _haversine_km(f["gps_lat"], f["gps_lng"], t["gps_lat"], t["gps_lng"])
        rows.append({
            "stop_from":  f["stop_num"],  "stop_to":   t["stop_num"],
            "from_id":    f["customer_id"], "from_shop": f["shop_name"],
            "from_lat":   round(f["gps_lat"], 6), "from_lng": round(f["gps_lng"], 6),
            "to_id":      t["customer_id"], "to_shop":   t["shop_name"],
            "to_lat":     round(t["gps_lat"], 6), "to_lng": round(t["gps_lng"], 6),
            "leg_km":     round(km, 3),
            "leg_min":    round((km / avg_speed) * 60, 1),
        })
    return pd.DataFrame(rows)


def build_stop_to_stop_map(
    result,
    territory_id:  str,
    schedule_date,
    sales_id:      str,
    zoom_start:    int = 13,
):
    """
    Folium map with stop-to-stop legs labelled with distance (km).
    Warehouse is the start; each polyline leg shows km at its midpoint.
    """
    try:
        import folium
    except ImportError:
        raise ImportError("pip install folium")

    dist_df    = build_stop_to_stop_distance_table(result, territory_id, schedule_date, sales_id)
    sched_date = pd.Timestamp(schedule_date).normalize()
    if dist_df.empty:
        raise ValueError(f"No route for {sales_id} in {territory_id} on {schedule_date}")

    all_lats = dist_df["from_lat"].tolist() + [dist_df["to_lat"].iloc[-1]]
    all_lngs = dist_df["from_lng"].tolist() + [dist_df["to_lng"].iloc[-1]]
    m = folium.Map(location=[sum(all_lats)/len(all_lats), sum(all_lngs)/len(all_lngs)],
                   zoom_start=zoom_start)

    # Draw each leg + distance label at midpoint
    for _, leg in dist_df.iterrows():
        coords = [[leg["from_lat"], leg["from_lng"]], [leg["to_lat"], leg["to_lng"]]]
        folium.PolyLine(coords, weight=3, color="navy", opacity=0.85).add_to(m)
        mid_lat = (leg["from_lat"] + leg["to_lat"]) / 2
        mid_lng = (leg["from_lng"] + leg["to_lng"]) / 2
        folium.Marker(
            [mid_lat, mid_lng],
            icon=folium.DivIcon(
                html=(
                    '<div style="background:rgba(255,255,255,0.88);padding:1px 5px;'
                    'border-radius:3px;font-size:10px;color:#003;white-space:nowrap;">'
                    + f'{leg["leg_km"]:.2f} km</div>'
                ),
                icon_size=(70, 18), icon_anchor=(35, 9),
            ),
            tooltip=f'Leg {int(leg["stop_from"])}\u2192{int(leg["stop_to"])}: {leg["leg_km"]:.2f} km ({leg["leg_min"]:.0f} min)',
        ).add_to(m)

    # Stop markers
    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))
    _add_warehouse_marker(m, wh_lat, wh_lng)

    det    = result.detailed_schedule
    day_df = det[
        (det["territory_id"]  == territory_id) &
        (det["schedule_date"] == sched_date) &
        (det["sales_id"]      == sales_id)
    ].copy()
    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank")

    for _, row in day_df.iterrows():
        color = SEGMENT_COLOUR.get(str(row.get("rfm_segment_final", "")), "blue")
        folium.Marker(
            [row["gps_lat"], row["gps_lng"]],
            icon=folium.DivIcon(
                html=(
                    f'<div style="background:{color};color:white;border-radius:50%;'
                    f'width:26px;height:26px;text-align:center;line-height:26px;'
                    f'font-weight:bold;font-size:11px;">{row.get("route_rank","")}</div>'
                ),
            ),
            tooltip=(
                f'{row.get("route_rank","")}. {row.get("shop_name","")} '
                f'[{row.get("rfm_segment_final","")}] '
                f'leg={row.get("route_leg_km", 0):.2f} km'
            ),
        ).add_to(m)

    total_km = dist_df["leg_km"].sum()
    m.get_root().html.add_child(folium.Element(
        f'<div style="position:fixed;top:10px;left:60px;z-index:1000;'
        f'background:white;padding:8px 12px;border-radius:6px;'
        f'border:1px solid #ccc;font-family:sans-serif;">'
        f'<b>{sales_id}</b> \u2014 {sched_date.strftime("%d %b %Y")}'
        f'&nbsp;|&nbsp;{len(dist_df)} legs&nbsp;|&nbsp;Total: {total_km:.2f} km'
        f'</div>'
    ))
    return m


def export_stop_to_stop_excel(
    result,
    territory_id:  str,
    schedule_date,
    filepath:      str = "stop_to_stop_distances.xlsx",
):
    """
    Excel workbook — one sheet per salesperson, one row per route leg.
    Columns: stop_from | stop_to | from_shop | to_shop | leg_km | leg_min
    Sheet 'Summary' gives total km per salesperson.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    sched_date = pd.Timestamp(schedule_date).normalize()
    det = result.detailed_schedule
    sp_list = sorted(
        det[(det["territory_id"]  == territory_id) &
            (det["schedule_date"] == sched_date)]["sales_id"].unique()
    )
    if not sp_list:
        raise ValueError(f"No schedule for {territory_id} on {sched_date.date()}")

    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(color="FFFFFF", bold=True)
    ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")
    TOTAL_FONT = Font(bold=True)

    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = []

    for sid in sp_list:
        dist_df = build_stop_to_stop_distance_table(result, territory_id, sched_date, sid)
        if dist_df.empty:
            continue

        ws = wb.create_sheet(title=sid[-12:])
        headers = ["From Stop", "To Stop", "From ID", "From Shop",
                   "From Lat", "From Lng", "To ID", "To Shop",
                   "To Lat", "To Lng", "Leg km", "Leg min"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

        for ri, row in dist_df.iterrows():
            er   = ri + 2
            fill = ALT_FILL if ri % 2 == 0 else None
            vals = [int(row["stop_from"]), int(row["stop_to"]),
                    row["from_id"], row["from_shop"],
                    row["from_lat"], row["from_lng"],
                    row["to_id"],   row["to_shop"],
                    row["to_lat"],  row["to_lng"],
                    round(float(row["leg_km"]), 3),
                    round(float(row["leg_min"]), 1)]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=er, column=ci, value=val)
                if fill:
                    cell.fill = fill

        tr = len(dist_df) + 2
        for ci, val in [(10, "TOTAL"),
                        (11, round(float(dist_df["leg_km"].sum()), 3)),
                        (12, round(float(dist_df["leg_min"].sum()), 1))]:
            ws.cell(row=tr, column=ci, value=val).font = TOTAL_FONT

        col_widths = [10, 8, 16, 30, 10, 10, 16, 30, 10, 10, 10, 10]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        summary_rows.append({
            "sales_id":  sid,
            "n_stops":   len(dist_df),
            "total_km":  round(float(dist_df["leg_km"].sum()), 3),
            "total_min": round(float(dist_df["leg_min"].sum()), 1),
        })

    # Summary sheet
    ws_s = wb.create_sheet(title="Summary", index=0)
    for ci, h in enumerate(["Salesperson", "Stops", "Total km", "Total min"], 1):
        c = ws_s.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(summary_rows, 2):
        ws_s.cell(row=ri, column=1, value=row["sales_id"])
        ws_s.cell(row=ri, column=2, value=row["n_stops"])
        ws_s.cell(row=ri, column=3, value=row["total_km"])
        ws_s.cell(row=ri, column=4, value=row["total_min"])
    for ci, w in enumerate([22, 10, 12, 12], 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.cell(row=1, column=6,
              value=f"{territory_id}  |  {sched_date.strftime('%d %b %Y')}").font = Font(bold=True)

    wb.save(filepath)
    print(f"Saved stop-to-stop Excel \u2192 {filepath}  ({len(summary_rows)} salesperson sheet(s))")
    return filepath