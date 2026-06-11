"""
At this point, we have
    1. Priority wise customer segment according to RFM
    2. Holidays
    3. Preferred day
    4. Cold or Warm truck req

Solver
Hard Constraint
    1. Driver daily working limit is 480min including travelling and serving
    2. Custonmer can get max visit as its segment allows

Soft Constraing
    1. Priority wise scheduling should be done, high then med then low
    2. Preferred day should get extra bonus for each customer
    3. Cusomter can get max 1 visit on each day

* Route is considered as one way wh-> cus1->cu2->....->cusn (one way only considered)
* min 1 visit is a filteration not a constraint

*Things assumed:
1. Travelling speeed is 32kmph
2. Serving time to each customer is 22min

3. Solver to 600s for warm else 300 for cold
"""

"""
Coding Steps

A. Schduling should be done territory wise
    1. Segregate cold/warm -> schduler solve fior them intevially
    2. Blcoked dates (customer holiday + salesman holiday)
    3. customer rfm segment (H/M/L)
    4. preference weight set
    5. Constraint add
        
        Hard
        5.1 work limit <=480min
        5.2 visit <= segment allows
        
        Soft
        5.3 Prefernce to high -> med
        5.4 add bonus for day match

B. Now consider the schedle created and use NN to route customer on same route to single salesman (Daily routing)
"""
import os
import numpy as np
import pandas as pd
import math
from typing import Optional, List, Dict, Tuple, Any
from dataclasses import *
import calendar

from ortools.sat.python import cp_model 
from ortools.constraint_solver import routing_enums_pb2
from ortools.constraint_solver import pywrapcp

SEG_CAPS = {"High": 20, "Medium": 11, "Low": 5}
DEFAULT_SEG_CAPS = {"High": 4, "Medium": 2, "Low": 1}
MIN_VISIT = 1
DEFAULT_DAILY_WORK_MINUTES = 480
DEFAULT_AVG_VISIT_MINUTES  = 22
DEFAULT_AVG_SPEED_KMPH     = 32
DEFAULT_SOLVER_TIME = 1200
MIN_SOLVER_TIME = 600
MAX_SOLVER_TIME = 1200

#Helper functions

# Calculate the exact distance (in kilometers) from each individual customer's store to those center points
def _haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R    = 6371.0088
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a    = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1 - a))

# where the center points (centroids) of the clusters are
def _kmeans_centroids(coords: np.ndarray, k: int, max_iter: int = 20) -> np.ndarray:
    if len(coords) <= k:
        return coords
    rng = np.random.default_rng(42)
    centroids = coords[rng.choice(len(coords), k, replace=False)]
    for _ in range(max_iter):
        dists = np.linalg.norm(coords[:, np.newaxis] - centroids, axis=2)
        labels = np.argmin(dists, axis=1)
        new_centroids = np.array([
            coords[labels == j].mean(axis=0) if np.any(labels == j) else centroids[j]
            for j in range(k)
        ])
        if np.allclose(centroids, new_centroids):
            break
        centroids = new_centroids
    return centroids

def visit_allowed_by_segment(segment: str) -> int:
    return SEG_CAPS.get(segment, MIN_VISIT)

def get_segment_priority_weight(segment: str) -> int:
    """
    High >> Medium >> Low in objective.

    The gap must be so large that ONE extra High visit is worth more than
    ALL possible Low visits combined.  This forces the solver to fully
    saturate High before touching Medium, and Medium before Low.

    With Low=10, 29 Low customers × 10 visits × 10 = 2,900  total Low value.
    One High visit = 10,000,000.  ⇒ strict priority.
    """
    return {"High": 10_00000, "Medium": 2000, "Low": 10}.get(segment, 10)

def _blocked_dates(
    col_name: str,
    col_value: str,
    holiday_df: pd.DataFrame,
    month_start: pd.Timestamp,
    month_end: pd.Timestamp,
) -> set[pd.Timestamp]:
    blocked: set[pd.Timestamp] = set()
    hdf = holiday_df.copy()
    hdf["from_date"] = pd.to_datetime(hdf["from_date"]).dt.normalize()
    hdf["to_date"]   = pd.to_datetime(hdf["to_date"]).dt.normalize()
    for _, row in hdf[hdf[col_name] == col_value].iterrows():
        cur = row["from_date"]
        while cur <= row["to_date"]:
            blocked.add(cur)
            cur += pd.Timedelta(days=1)
    return {d for d in blocked if month_start <= d <= month_end}


def get_salesperson_blocked_dates(sales_id, holiday_df, month_start, month_end):
    return _blocked_dates("salesperson_holiday", sales_id, holiday_df, month_start, month_end)

def get_territory_blocked_dates(territory_id, holiday_df, month_start, month_end):
    return _blocked_dates("territory_holiday", territory_id, holiday_df, month_start, month_end)

def _build_full_customer_with_rfm_score_df(customer_df: pd.DataFrame, rfm_scores_df: pd.DataFrame) -> pd.DataFrame:
    rfm_cols = [
        "customer_id", "rfm_segment_final", "final_customer_score",
        "rfm_combined", "customer_rank", "recency", "frequency", "monetary",
        "r_score", "f_score", "m_score",
        "seasonality_score", "territory_score", "locality_score", "rating_score",
    ]
    rfm_use = [c for c in rfm_cols if c in rfm_scores_df.columns]
    merged  = customer_df.merge(rfm_scores_df[rfm_use], on="customer_id", how="left")
    merged["rfm_segment_final"]    = merged["rfm_segment_final"].fillna("Low")
    merged["final_customer_score"] = merged["final_customer_score"].fillna(0.0)
    return merged

# ─────────────────────────────────────────────────────────────────────────────
# Solver failure diagnostics
# ─────────────────────────────────────────────────────────────────────────────
def _diagnose_infeasible(
    customer_ids: list[str],
    salesperson_ids: list[str],
    valid_dates: list[pd.Timestamp],
    cust_lookup: dict,
    sp_is_cold: dict[str, bool],
    avg_visit_min: int,
    travel_mins: dict[tuple[str, str], int],     # (cid, sid) → travel minutes one-way
    daily_work_min: int,
    truck_group: str,                              # "cold" or "normal"
) -> list[str]:
    """
    Re-solves relaxed versions of the model to identify which constraint
    class is causing infeasibility.  Returns a list of human-readable
    diagnostic strings.
    """
    diagnostics: list[str] = []
    n_cust = len(customer_ids)
    n_sp   = len(salesperson_ids)
    n_days = len(valid_dates)

    # --- capacity arithmetic check (no CP model needed) ----------------------
    # Worst case: every customer assigned to the most distant SP on the same day
    max_per_sp_day: dict[str, int] = {}
    for sid in salesperson_ids:
        avg_t = int(np.mean([travel_mins.get((cid, sid), 0) for cid in customer_ids]) or 0)
        cap   = max(1, daily_work_min // max(avg_visit_min + avg_t, 1))
        max_per_sp_day[sid] = cap

    total_monthly_capacity = sum(
        max_per_sp_day[sid] * n_days for sid in salesperson_ids
    )

    if n_cust > total_monthly_capacity:
        diagnostics.append(
            f"CAPACITY: {n_cust} customers cannot be scheduled with {n_sp} salesperson(s) "
            f"across {n_days} working days.  Total monthly capacity = {total_monthly_capacity} slots "
            f"(based on avg travel times and {daily_work_min} min/day).  "
            f"Fix: reduce customers per salesperson, add salespeople, or extend working hours."
        )

    # --- cold-truck check ----------------------------------------------------
    if truck_group == "cold":
        cold_sp = [s for s in salesperson_ids if sp_is_cold.get(s, False)]
        if not cold_sp:
            diagnostics.append(
                "COLD TRUCK: No cold-capable salesperson is available in this territory "
                "for the cold-truck customer group.  Assign a cold-van to at least one "
                "salesperson in the territory."
            )

    # --- working days check --------------------------------------------------
    if n_days == 0:
        diagnostics.append(
            "HOLIDAYS: All working days in the month are blocked by territory or "
            "salesperson holidays.  No valid dates remain to schedule visits."
        )

    # --- per-SP one-visit-per-day limit check --------------------------------
    worst_sp_cap = max(max_per_sp_day.values()) if max_per_sp_day else 0
    if n_cust > worst_sp_cap * n_days * n_sp:
        diagnostics.append(
            f"DAILY LIMIT: Even with all salespeople working every day, the daily "
            f"capacity ({worst_sp_cap} customers/SP/day) × {n_days} days × {n_sp} SPs "
            f"= {worst_sp_cap * n_days * n_sp} total slots is less than {n_cust} customers."
        )

    # --- min-1-visit hard constraint check -----------------------------------
    # This can be infeasible if a customer's only valid date is also at capacity.
    # Approximation: if total slots < n_customers, min-visit constraint is infeasible.
    if total_monthly_capacity < n_cust and not diagnostics:
        diagnostics.append(
            "MIN VISIT: The hard constraint 'every customer ≥ 1 visit' cannot be satisfied "
            f"because total available slots ({total_monthly_capacity}) < customers ({n_cust}).  "
            "Reduce the customer pool or increase salesperson capacity."
        )

    if not diagnostics:
        diagnostics.append(
            "UNKNOWN: CP-SAT returned INFEASIBLE but no arithmetic constraint breach was "
            "detected.  This may be due to a combination of holiday blocking, cold-truck "
            "restrictions, and tight capacity. Try increasing solver_time_seconds or "
            "enabling relaxed_min_visits=True."
        )

    return diagnostics


def _report_failure_diagnostics(
    solver: cp_model.CpSolver,
    status: int,
    territory_id: int,
    truck_group: str,
    visit: dict,
    assigned: dict,
    customer_ids: list[str],
    sp_ids: list[str],
    valid_dates: list[pd.Timestamp],
    cust_lookup: dict,
    sp_is_cold: dict[str, bool],
    avg_visit_minutes: int,
    travel_mins: dict[tuple[str, str], int],
    daily_work_minutes: int,
) -> None:
    status_name = solver.StatusName(status)
    print(f"\n  ❌ [{territory_id}/{truck_group}] CP-SAT returned: {status_name}")
    print(f"     Variables : {len(visit) + len(assigned)}")
    print(f"     Customers : {len(customer_ids)}")
    print(f"     Salespeople: {len(sp_ids)}")
    print(f"     Valid dates: {len(valid_dates)}")
    diagnostics = _diagnose_infeasible(
        customer_ids, sp_ids, valid_dates, cust_lookup,
        sp_is_cold, avg_visit_minutes, travel_mins, daily_work_minutes, truck_group,
    )
    print(f"\n  ── Constraint failure analysis ──")
    for i, msg in enumerate(diagnostics, 1):
        print(f"     [{i}] {msg}")
    print()


#Data Classes
@dataclass
class SalespersonScheduleResult:
    sales_id:           str
    territory_id:       str
    detailed_schedule:  pd.DataFrame
    daily_schedule:     pd.DataFrame
    assigned_customers: pd.DataFrame

@dataclass
class MultiScheduleResult:
    detailed_schedule:    pd.DataFrame   # all customers, all territories
    cold_schedule:        pd.DataFrame   # cold-truck customers only
    normal_schedule:      pd.DataFrame   # normal-truck customers only
    daily_schedule:       pd.DataFrame
    unvisited_customers:     pd.DataFrame   # customers who received 0 visits (capacity overflow)
    under_visited_customers: pd.DataFrame   # customers with >0 but < segment cap visits
    daily_visit_plan:        pd.DataFrame   # compact: sales_id | schedule_date | customer_visits (list of dicts)
    salesperson_results:     dict[str, SalespersonScheduleResult] = field(default_factory=dict)
    territory_warehouses:    dict[str, tuple[float, float]]       = field(default_factory=dict)



class MultiSalesManScheduler:
    def __init__(self, config_df: dict | pd.DataFrame):
        if isinstance(config_df, pd.DataFrame):
            cfg = (
                config_df.set_index("config_key")["config_value"].to_dict()
                if not config_df.empty else {}
            )
        else:
            cfg = config_df if config_df else {}
        self.avg_speed = float(cfg.get('avg_speed', 32))
        self.customer_serving_time = int(cfg.get('customer_serving_time', 22))
        self.salesman_daily_work_minutes = int(cfg.get('salesman_daily_work_minutes', 480))
        self._territory_scheduler = TerritoryScheduler()
        self._route_planner = DailyRoutePlanner()

    def create_monthly_schedule(
        self,
        customer_df:      pd.DataFrame,
        rfm_scores_df:    pd.DataFrame,
        salesperson_df:   pd.DataFrame,
        holiday_df:       pd.DataFrame,
        territory_df:     pd.DataFrame,
        van_df:           pd.DataFrame,
        month_start_date: str | pd.Timestamp = "2026-07-01",
        territory_id:     Optional[str]      = None,   # None → all territories
    ) -> MultiScheduleResult:
        
        if (territory_df is None or territory_df.empty or
            customer_df is None or customer_df.empty or
            salesperson_df is None or salesperson_df.empty):
            return MultiScheduleResult()

        month_start   = pd.Timestamp(month_start_date).normalize()
        year, month   = month_start.year, month_start.month
        days_in_month = calendar.monthrange(year, month)[1]
        month_end     = month_start + pd.Timedelta(days=days_in_month - 1)

        customers_with_rfm = _build_full_customer_with_rfm_score_df(customer_df, rfm_scores_df)
        territory_info = territory_df.set_index("territory_id").to_dict('index')

        generate_plan_for_territory = territory_df[territory_df["territory_id"] == territory_id] if territory_id else territory_df 

        all_detailed: list[pd.DataFrame] = []
        sp_results:     dict[str, SalespersonScheduleResult] = {}

        for _, territory_row in generate_plan_for_territory.iterrows():
            # Extracting out the territory ID, customers and the salesman
            terr_id = territory_row["territory_id"]
            terr_customers = customers_with_rfm[customers_with_rfm["territory_id"] == terr_id].copy()
            terr_salesman = salesperson_df[salesperson_df["territory_id"] == terr_id].copy() 
        
            # Chcking if the customers or salesman DF is empty             
            if terr_customers.empty or terr_salesman.empty:
                print(f"""For Territory - {terr_id}:
                      Customers: {len(terr_customers)}
                      Salesman: {len(terr_salesman)}
                    """)
                continue
            
            #Extracting the territory's warehouse latitute and longitude
            terr_warehouse_lat = territory_info[terr_id]['warehouse_lat']
            terr_warehouse_lng = territory_info[terr_id]['warehouse_lng']

            #Territory blocked dates
            terr_blocked_dates = get_territory_blocked_dates(terr_id, holiday_df,month_start, month_end)
            print(f"Only Territory {terr_id} working days: {days_in_month-len(terr_blocked_dates)} ")

            #Salesman blocked dates
            #Iterating on territory salesman's DF, to get salesman valid dates
            group_salesman_workings_dates: dict[str, list[pd.Timestamp]] = {}
            for _, salesman_row in terr_salesman.iterrows():
                terr_salesman_id = salesman_row['sales_id']
                salesman_blocked_dates = get_salesperson_blocked_dates(terr_salesman_id, holiday_df,month_start, month_end)
                print(f"Only Salesman {terr_salesman_id} working days: {days_in_month-len(salesman_blocked_dates)} ")
                
                all_blocked_dates = terr_blocked_dates | salesman_blocked_dates 
                
                group_salesman_workings_dates[terr_salesman_id] = [month_start + pd.Timedelta(days=i)
                                            for i in range(days_in_month)
                                            if (month_start + pd.Timedelta(days=i)) not in all_blocked_dates]
            
            print(f"Territory all salesman final working days: {group_salesman_workings_dates}")
            
            #Segregating cold and normal customers
            cold_customers = terr_customers[terr_customers["cold_truck_required"]== True].copy()
            normal_cusomters = terr_customers[terr_customers["cold_truck_required"]== False].copy()

            for group_name, group_customers in [("cold", cold_customers), ("normal", normal_cusomters)]:
                if group_customers.empty:
                    print(f" Territory {terr_id} has  NO {group_name} customers")
                    continue

                #Determine active salesperson and assigned to there respective group
                # Cold SPs (those with a cold-enabled van) are exclusively
                # reserved for cold-truck customers.  Normal-truck customers
                # are served only by the remaining SPs.
                # This ensures a SP dedicated to cold routes is never
                # double-counted in the normal pool.`
                van_cold_map = van_df.set_index("van_id")["cold_truck_enabled"].to_dict()
                
                
                cold_sp_mask = terr_salesman["assigned_van"].map(van_cold_map).fillna(False)
                cold_salesman = terr_salesman[cold_sp_mask].copy()
                normal_salesman   = terr_salesman[~cold_sp_mask].copy()
                
                if group_name == "cold":
                    group_salesman = cold_salesman 
                    if cold_salesman.empty:
                        print(f"Territory {terr_id} has NO cold salesman. All SPs are normal-capable — sharing full pool as fallback.")
                        group_salesman = terr_salesman.copy()
                
                else:
                    group_salesman = normal_salesman
                    if normal_salesman.empty:
                        print(f"Territory {terr_id} has NO noraml salesman. All SPs are cold-capable — sharing full pool as fallback.")
                        group_salesman = terr_salesman.copy()
                
                #Now we have salesman truck group, and final working dates
                active_salesman_working_dates = {
                    salesman_id: group_salesman_workings_dates[salesman_id] 
                    for salesman_id in group_salesman["sales_id"].to_list()
                    if salesman_id in group_salesman_workings_dates
                }
                
                group_salesman_all_valid_dates = {
                    d for dates in active_salesman_working_dates.values() for d in dates
                }
                
                datailed_schedule = self._territory_scheduler.solve(
                    #territory group related params
                    territory_id = terr_id,
                    customer_df = group_customers,
                    salesman_df = group_salesman,
                    truck_category = group_name,
                    van_df = van_df,
                    valid_dates = group_salesman_all_valid_dates,
                    group_salesman_workings_dates = active_salesman_working_dates,
                    
                    terr_warehouse_lat = terr_warehouse_lat,
                    terr_warehouse_lng = terr_warehouse_lng,
                    
                    #config set
                    avg_speed = self.avg_speed,
                    customer_serving_time = self.customer_serving_time,
                    salesman_daily_work_minutes = self.salesman_daily_work_minutes,                    
                    solver_time = MIN_SOLVER_TIME if group_name == 'cold' else MAX_SOLVER_TIME
                )
                
                if datailed_schedule.empty:
                    print("No plan generated!")
                    continue
                
                #Apply NN for best route and km traking
                # ── Post-process: nearest-neighbour routes + km tracking ──────
                datailed_schedule = self._apply_route_ordering(
                    datailed_schedule, terr_warehouse_lat, terr_warehouse_lng)
                
                # ── Greedy top-up: fill remaining daily capacity ──────────────
                detailed = self._greedy_topup(
                    detailed        = datailed_schedule,
                    group_custs     = group_customers,
                    valid_dates     = group_salesman_all_valid_dates,
                    wh_lat          = terr_warehouse_lat,
                    wh_lng          = terr_warehouse_lng,
                    avg_visit_minutes  = self.customer_serving_time,
                    avg_speed_kmph     = self.avg_speed,
                    daily_work_minutes = self.salesman_daily_work_minutes,
                    territory_id    = terr_id,
                    truck_group     = group_name,
                    sp_valid_dates  = group_salesman_workings_dates,
                )
                
                all_detailed.append(detailed)

                # Store per-salesperson results
                for salesman_id in detailed["sales_id"].unique():
                    sp_det = detailed[detailed["sales_id"] == salesman_id].copy()
                    sp_cust= group_customers[group_customers["customer_id"].isin(
                        sp_det["customer_id"].unique()
                    )].copy()
                    daily  = self._build_daily_summary(sp_det, terr_warehouse_lat, terr_warehouse_lng, salesman_id)

                    key = f"{salesman_id}_{group_name}"   # unique key per SP × truck group
                    sp_results[key] = SalespersonScheduleResult(
                        sales_id           = salesman_id,
                        territory_id       = terr_id,
                        detailed_schedule  = sp_det,
                        daily_schedule     = daily,
                        assigned_customers = sp_cust,
                    )


        # ── Combine all results ───────────────────────────────────────────────
        if all_detailed:
            combined = (
                pd.concat(all_detailed, ignore_index=True)
                .sort_values(
                    ["territory_id", "truck_group", "sales_id",
                     "schedule_date", "final_customer_score"],
                    ascending=[True, True, True, True, False],
                )
                .reset_index(drop=True)
            )
        else:
            combined = TerritoryScheduler._empty_df()

        cold_schedule   = combined[combined["truck_group"] == "cold"].reset_index(drop=True)
        normal_schedule = combined[combined["truck_group"] == "normal"].reset_index(drop=True)
        combined_daily  = self._build_combined_daily(sp_results)
        ter_warehouses: dict[str, tuple[float, float]] = {}

        _CAPS = SEG_CAPS

        scheduled_ids = set(combined["customer_id"].unique()) if not combined.empty else set()

        # Scope to the requested territory if applicable
        scoped_customers = customers_with_rfm.copy()
        if territory_id:
            scoped_customers = scoped_customers[
                scoped_customers["territory_id"] == territory_id
            ]

        # ── Bucket 1: Unvisited — 0 visits ────────────────────────────────────
        unvisited_df = (
            scoped_customers[~scoped_customers["customer_id"].isin(scheduled_ids)]
            .copy()
            .reset_index(drop=True)
        )
        if not unvisited_df.empty:
            print(f"\n  ⚠️  {len(unvisited_df)} customer(s) received NO visits "
                  f"(capacity overflow — not a solver error):")
            for _, r in unvisited_df.iterrows():
                print(f"     {r['customer_id']}  [{r.get('rfm_segment_final','?')}]  "
                      f"{r.get('shop_name','')}  ({r.get('territory_id','')})")
        else:
            print(f"\n  ✅  All customers received at least 1 visit.")

        # ── Bucket 2: Under-visited — >0 but < segment cap ───────────────────
        if not combined.empty:
            visit_counts = (
                combined.groupby("customer_id")
                .size()
                .rename("actual_visits")
                .reset_index()
            )
            under_base = scoped_customers.merge(visit_counts, on="customer_id", how="left")
            under_base["actual_visits"] = under_base["actual_visits"].fillna(0).astype(int)
            
            visit_caps_map = combined.groupby("customer_id")["visit_cap"].first().to_dict()
            under_base["visit_cap"] = (
                under_base["customer_id"]
                .map(visit_caps_map)
                .fillna(under_base["rfm_segment_final"].map(_CAPS))
                .astype(int)
            )
            under_base["visits_gap"] = under_base["visit_cap"] - under_base["actual_visits"]
            under_base["pct_of_cap"] = (
                under_base["actual_visits"] / under_base["visit_cap"] * 100
            ).round(1)
            under_df = (
                under_base[
                    (under_base["actual_visits"] > 0) &
                    (under_base["visits_gap"]    > 0)
                ]
                .sort_values(
                    ["rfm_segment_final", "visits_gap"],
                    ascending=[True, False],
                )
                .reset_index(drop=True)
            )
        else:
            under_df = pd.DataFrame()

        if not under_df.empty:
            print(f"\n  ⚠️  {len(under_df)} customer(s) under-visited (>0 but < cap):")
            for seg, grp in under_df.groupby("rfm_segment_final"):
                cap_val = grp['visit_cap'].iloc[0]
                print(f"     {seg}: {len(grp)} customers | "
                      f"avg {grp['actual_visits'].mean():.1f}/{cap_val} visits | "
                      f"total missed = {grp['visits_gap'].sum()}")
        else:
            print(f"\n  ✅  All scheduled customers reached their full visit cap.")

        # ── Build compact daily visit plan ───────────────────────────────────
        daily_visit_plan = self._build_daily_visit_plan(combined)
        ter_warehouses[terr_id] = (terr_warehouse_lat, terr_warehouse_lng)
        return MultiScheduleResult(
            detailed_schedule       = combined,
            cold_schedule           = cold_schedule,
            normal_schedule         = normal_schedule,
            daily_schedule          = combined_daily,
            unvisited_customers     = unvisited_df,
            under_visited_customers = under_df,
            daily_visit_plan        = daily_visit_plan,
            salesperson_results     = sp_results,
            territory_warehouses    = ter_warehouses,
        )

    def _apply_route_ordering(self, datailed_schedule, terr_warehouse_lat, terr_warehouse_lng) -> pd.DataFrame:
        """
            For each (sales_id, schedule_date), apply nearest-neighbour ordering
            from the warehouse and populate route_leg_km / cumulative_route_km.
        """
        parts: list[pd.DataFrame] = []
        for (sid, d), group in datailed_schedule.groupby(["sales_id", "schedule_date"]):
            routed = self._route_planner.get_route(group.copy(), terr_warehouse_lat, terr_warehouse_lng)
            parts.append(routed)
        if not parts: return datailed_schedule
        return pd.concat(parts, ignore_index=True)

    def _greedy_topup(
        self,
        detailed: pd.DataFrame,
        group_custs: pd.DataFrame,
        valid_dates: list[pd.Timestamp],
        wh_lat: float, wh_lng: float,
        avg_visit_minutes: int,
        avg_speed_kmph: float,
        daily_work_minutes: int,
        territory_id: str,
        truck_group: str,
        sp_valid_dates: dict[str, list[pd.Timestamp]] | None = None,
    ) -> pd.DataFrame:
        """
        Post-solve greedy fill: for each (SP, day) with remaining time budget,
        insert extra visits for under-visited customers assigned to that SP.

        Uses the real NN-routed distances (route_leg_km must be populated)
        to compute current time usage, then tries inserting the highest-priority
        under-visited customers that fit in the remaining minutes.
        """
        if detailed.empty:
            return detailed

        _CAPS = SEG_CAPS
        if not detailed.empty and "fallback_used" in detailed.columns:
            if detailed["fallback_used"].iloc[0]:
                _CAPS = DEFAULT_SEG_CAPS
        cust_lookup = group_custs.set_index("customer_id").to_dict("index")

        # Customer → assigned SP (from solver output)
        cust_sp = detailed.groupby("customer_id")["sales_id"].first().to_dict()

        # For customers not in solver output, assign to the closest salesperson in this group
        sp_coords = {}
        for sid, grp in detailed.groupby("sales_id"):
            sp_coords[sid] = (grp["gps_lat"].mean(), grp["gps_lng"].mean())

        for cid in group_custs["customer_id"]:
            if cid not in cust_sp:
                # Find closest salesperson in this group based on average scheduled coordinates
                c_lat = float(cust_lookup[cid]["gps_lat"])
                c_lng = float(cust_lookup[cid]["gps_lng"])
                best_sid = None
                min_d = float('inf')
                for sid, coords in sp_coords.items():
                    d = _haversine_km(c_lat, c_lng, coords[0], coords[1])
                    if d < min_d:
                        min_d = d
                        best_sid = sid
                # If salesperson average coordinates are not available (e.g. no visits scheduled yet),
                # assign to the first salesperson in the group
                if best_sid is None:
                    active_sps = list(sp_coords.keys())
                    if active_sps:
                        best_sid = active_sps[0]
                if best_sid is not None:
                    cust_sp[cid] = best_sid

        # Current visit count per customer (including solver output only)
        visit_counts = detailed.groupby("customer_id").size().to_dict()

        # Build per-SP list of under-visited candidates, sorted by priority
        sp_candidates: dict[str, list[str]] = {}
        for cid, sid in cust_sp.items():
            if cid not in cust_lookup:
                continue
            seg    = cust_lookup[cid].get("rfm_segment_final", "Low")
            cap    = _CAPS.get(seg, 10)
            actual = visit_counts.get(cid, 0)
            if actual < cap:
                sp_candidates.setdefault(sid, []).append(cid)

        for sid in sp_candidates:
            sp_candidates[sid].sort(key=lambda c: (
                -get_segment_priority_weight(cust_lookup[c].get("rfm_segment_final", "Low")),
                -cust_lookup[c].get("final_customer_score", 0.0),
            ))

        if not sp_candidates:
            return detailed

        # Track additional visits added (so we don't exceed cap)
        added: dict[str, int] = {}
        new_rows: list[dict] = []

        for sid in detailed["sales_id"].unique():
            candidates = sp_candidates.get(sid, [])
            if not candidates:
                continue

            sp_dates = set(sp_valid_dates.get(sid, valid_dates)) if sp_valid_dates else set(valid_dates)

            for d in sorted(valid_dates):
                if d not in sp_dates:
                    continue

                day_df = detailed[
                    (detailed["sales_id"] == sid) &
                    (detailed["schedule_date"] == d)
                ]
                if day_df.empty:
                    # Solver gave this SP no visits today — start fresh from warehouse
                    remaining    = float(daily_work_minutes)
                    cur_lat, cur_lng = wh_lat, wh_lng
                    already_today: set[str] = set()
                else:
                    # Compute current time usage from real routed distances
                    route_km    = float(day_df["route_leg_km"].sum()) if "route_leg_km" in day_df.columns else 0.0
                    travel_min  = (route_km / avg_speed_kmph) * 60
                    service_min = len(day_df) * avg_visit_minutes
                    remaining   = daily_work_minutes - (travel_min + service_min)
                    
                    print(f"remaining time for {sid} is {remaining}")
                    print(f"Service min for {sid}: {service_min}")
                    print(f"Travel min for {sid}: {travel_min}")

                    # Last stop coords for cheapest-insertion estimate
                    if "route_rank" in day_df.columns:
                        last = day_df.sort_values("route_rank").iloc[-1]
                    else:
                        last = day_df.iloc[-1]
                    cur_lat, cur_lng = float(last["gps_lat"]), float(last["gps_lng"])
                    already_today = set(day_df["customer_id"].tolist())

                if remaining < avg_visit_minutes + 2:
                    continue

                for cid in candidates:
                    if cid in already_today:
                        continue
                    total_visits = visit_counts.get(cid, 0) + added.get(cid, 0)
                    seg = cust_lookup[cid].get("rfm_segment_final", "Low")
                    if total_visits >= _CAPS.get(seg, 10):
                        continue

                    c_lat = float(cust_lookup[cid]["gps_lat"])
                    c_lng = float(cust_lookup[cid]["gps_lng"])
                    leg_km         = _haversine_km(cur_lat, cur_lng, c_lat, c_lng)
                    leg_travel_min = (leg_km / avg_speed_kmph) * 60
                    cost           = avg_visit_minutes + leg_travel_min

                    if cost > remaining:
                        continue

                    # Insert this visit
                    info = cust_lookup[cid]
                    wh_km = _haversine_km(wh_lat, wh_lng, c_lat, c_lng)
                    new_rows.append({
                        "schedule_date":            d,
                        "sales_id":                 sid,
                        "territory_id":             territory_id,
                        "customer_id":              cid,
                        "truck_group":              truck_group,
                        "shop_name":                info.get("shop_name", ""),
                        "locality":                 info.get("locality", ""),
                        "gps_lat":                  info.get("gps_lat"),
                        "gps_lng":                  info.get("gps_lng"),
                        "rfm_segment_final":        info.get("rfm_segment_final", "Low"),
                        "final_customer_score":     info.get("final_customer_score", 0.0),
                        "rfm_combined":             info.get("rfm_combined", 0.0),
                        "customer_rank":            info.get("customer_rank", 0),
                        "seasonality_score":        info.get("seasonality_score", 0.0),
                        "territory_score":          info.get("territory_score", 0.0),
                        "locality_score":           info.get("locality_score", 0.0),
                        "rating_score":             info.get("rating_score", 0.0),
                        "lifecycle_state":          info.get("lifecycle_state", "Active"),
                        "cold_truck_required":      info.get("cold_truck_required", False),
                        "estimated_visit_minutes":  avg_visit_minutes,
                        "estimated_travel_minutes": round(leg_travel_min, 1),
                        "warehouse_to_customer_km": round(wh_km, 3),
                        "route_leg_km":             None,
                        "cumulative_route_km":      None,
                        "visit_cap":                _CAPS.get(seg, 10),
                        "fallback_used":            not (_CAPS is SEG_CAPS),
                    })
                    remaining  -= cost
                    already_today.add(cid)
                    added[cid]  = added.get(cid, 0) + 1
                    cur_lat, cur_lng = c_lat, c_lng

                    if remaining < avg_visit_minutes + 2:
                        break

        if new_rows:
            n_added = len(new_rows)
            new_df  = pd.DataFrame(new_rows)
            detailed = pd.concat([detailed, new_df], ignore_index=True)
            # Re-run NN route ordering so route_leg_km is correct with new stops
            detailed = self._apply_route_ordering(detailed, wh_lat, wh_lng)
            print(f"  📈 Territory - {territory_id} of group - {truck_group} -  Greedy top-up inserted "
                  f"{n_added} extra visit(s) across {len(set(r['customer_id'] for r in new_rows))} customer(s)")
        return detailed

    def _build_daily_summary(
        self,
        detailed: pd.DataFrame,
        wh_lat: float,
        wh_lng: float,
        sales_id: str,
    ) -> pd.DataFrame:
        rows: list[dict] = []
        for d, group in detailed.groupby("schedule_date"):
            routed = group.sort_values("route_rank") if "route_rank" in group.columns else group
            total_leg_km = routed["route_leg_km"].sum() if "route_leg_km" in routed.columns else 0.0
            rows.append({
                "schedule_date":       d,
                "sales_id":            sales_id,
                "territory_id":        group["territory_id"].iloc[0],
                "truck_group":         group["truck_group"].iloc[0],
                "customer_list":       routed["customer_id"].tolist(),
                "customer_count":      len(routed),
                "route_order":         routed["shop_name"].tolist(),
                "segment_breakdown":   routed["rfm_segment_final"].value_counts().to_dict(),
                "avg_customer_score":  round(float(routed["final_customer_score"].mean()), 4),
                "total_visit_min":     int(routed["estimated_visit_minutes"].sum()),
                "total_travel_min": int((total_leg_km / self.avg_speed) * 60),
                "total_route_km":      round(float(total_leg_km), 3),
                "customer_km_detail":  dict(zip(
                    routed["customer_id"].tolist(),
                    routed["route_leg_km"].round(3).tolist()
                    if "route_leg_km" in routed.columns else [None] * len(routed),
                )),
            })
        if not rows:
            return pd.DataFrame(columns=[
                "schedule_date", "sales_id", "territory_id", "truck_group",
                "customer_list", "customer_count", "route_order", "segment_breakdown",
                "avg_customer_score", "total_visit_min", "total_travel_min",
                "total_route_km", "customer_km_detail",
            ])
        return pd.DataFrame(rows).sort_values(["sales_id", "schedule_date"]).reset_index(drop=True)

    @staticmethod
    def _build_combined_daily(sp_results: dict) -> pd.DataFrame:
        parts = [r.daily_schedule for r in sp_results.values() if not r.daily_schedule.empty]
        if not parts:
            return pd.DataFrame()
        return (
            pd.concat(parts, ignore_index=True)
            .sort_values(["territory_id", "truck_group", "sales_id", "schedule_date"])
            .reset_index(drop=True)
        )

    @staticmethod
    def _build_daily_visit_plan(combined: pd.DataFrame) -> pd.DataFrame:
        """
        Build a compact daily visit plan from the detailed schedule.

        Returns a DataFrame with columns:
            territory_id | sales_id | schedule_date | truck_group | customers
        where `customers` is a list of dicts:
            [{"customer_id": ..., "shop_name": ..., "rfm_segment": ..., "route_rank": ...}, ...]
        ordered by route_rank (nearest-neighbour from warehouse).
        """
        if combined.empty:
            return pd.DataFrame(columns=[
                "territory_id", "sales_id", "schedule_date", "truck_group", "customers",
            ])

        if "route_rank" in combined.columns:
            combined = combined.sort_values(
                ["territory_id", "sales_id", "schedule_date", "route_rank"]
            )

        rows: list[dict] = []
        group_keys = ["territory_id", "sales_id", "schedule_date", "truck_group"]
        for keys, grp in combined.groupby(group_keys):
            tid, sid, d, tg = keys
            if "route_rank" in grp.columns:
                grp = grp.sort_values("route_rank")
            customers = [
                {
                    "route_rank":   int(r.get("route_rank", 0)) if pd.notna(r.get("route_rank")) else i + 1,
                    "customer_id":  r["customer_id"],
                    "shop_name":    r.get("shop_name", ""),
                    "rfm_segment":  r.get("rfm_segment_final", ""),
                }
                for i, (_, r) in enumerate(grp.iterrows())
            ]
            rows.append({
                "territory_id":  tid,
                "sales_id":      sid,
                "schedule_date": pd.Timestamp(d).strftime("%Y-%m-%d"),
                "truck_group":   tg,
                "customers":     customers,
            })

        return (
            pd.DataFrame(rows)
            .sort_values(["territory_id", "sales_id", "schedule_date"])
            .reset_index(drop=True)
        )

class TerritoryScheduler:    
    def __init__(self): pass
    
    def solve(self,
        
        territory_id: str,
        customer_df: pd.DataFrame,
        salesman_df: pd.DataFrame,
        truck_category: str,
        van_df: pd.DataFrame,
        valid_dates: List[pd.Timestamp],
        group_salesman_workings_dates: Dict[str, List[pd.Timestamp] | Any],
        terr_warehouse_lat: float,
        terr_warehouse_lng: float,
        
        #config set
        avg_speed: int = DEFAULT_AVG_SPEED_KMPH,
        customer_serving_time: int = DEFAULT_AVG_VISIT_MINUTES,
        salesman_daily_work_minutes: int = DEFAULT_DAILY_WORK_MINUTES,
        
        solver_time: Optional[int] = DEFAULT_SOLVER_TIME
        ) -> pd.DataFrame:
        """
            Returns a detailed_schedule DataFrame (one row per customer-visit).
            Returns empty DataFrame on failure (with printed diagnostics).
        """
        model = cp_model.CpModel()
        customer_ids = customer_df['customer_id'].tolist()
        salesman_ids = salesman_df['sales_id'].tolist()
        
        #to return if plan not generated
        empty_df = self._empty_df()
        
        #check if customer id or salesman id not got
        
        if not customer_ids or not salesman_ids or not valid_dates:
            print(f"""
                Territory - {territory_id} of group  - {truck_category} has Nothing to schedule —
                Length of customers - {len(customer_ids)}
                Salesman - {len(salesman_ids)}
                valid dates - {len(valid_dates)}""")
            return empty_df
        
        #Cold Van capability Mapping
        van_cold = van_df.set_index("van_id")["cold_truck_enabled"].to_dict()
        
        salesman_df['is_cold'] = salesman_df["assigned_van"].map(van_cold).fillna(False)
        salesman_with_cold_truck: Dict[str, bool] = (
            salesman_df.groupby("sales_id")["is_cold"].any().to_dict()
            )
        
        customer_lookup_dict = customer_df.set_index("customer_id").to_dict("index")
        
        # ── Calculate direct travel times
        # Calculate warehouse-to-customer travel times
        _wh_legs = []
        for customer_id in customer_ids:
            c = customer_lookup_dict[customer_id]
            km = _haversine_km(terr_warehouse_lat, terr_warehouse_lng, float(c["gps_lat"]), float(c["gps_lng"]))
            _wh_legs.append(max(1, round(((km / avg_speed) * 60))))
        
        # Calculate all pairwise inter-customer travel times
        _all_inter = []
        for i, cid1 in enumerate(customer_ids):
            c1 = customer_lookup_dict[cid1]
            lat1, lng1 = float(c1["gps_lat"]), float(c1["gps_lng"])
            for j, cid2 in enumerate(customer_ids):
                if i == j:
                    continue
                c2 = customer_lookup_dict[cid2]
                lat2, lng2 = float(c2["gps_lat"]), float(c2["gps_lng"])
                km = _haversine_km(lat1, lng1, lat2, lng2)
                _all_inter.append(max(1, round(km / avg_speed * 60)))        

        # Scale down the raw average pairwise distance to account for geographic clustering of sequential stops (0.35 is geographically normalization is correct)
        avg_leg_travel_min = max(1, round(float(np.mean(_all_inter)) * 0.35)) if _all_inter else 4
        wh_leg_min = max(0, round(float(np.mean(_wh_legs)))) if _wh_legs else 5
        
        print(f"""
            Territory - {territory_id} of group - {truck_category}] averages calculated:
            No of Customers - {len(customer_ids)}
            Average Warehouse to customer distance (avg_wh_leg) - {wh_leg_min} minutes,
            Average customer to customer distance (avg_inter) - {avg_leg_travel_min} minutes
            Daily limit of salesman - {salesman_daily_work_minutes}
            """)
        
        #For printing puropose calculating Avg salesman's time per visit, and the daily customers the salesman can serve
        sp_time_visit: Dict[str, int] = {
            salesman_id: max(1, customer_serving_time + avg_leg_travel_min) for salesman_id in salesman_ids
        }
        sp_daily_visits: Dict[str, int] = {
            salesman_id: max(1, salesman_daily_work_minutes // sp_time_visit[salesman_id]) for salesman_id in salesman_ids
        }
        
        print(f"""Territory - {territory_id} of group - {truck_category}
            {list(sp_daily_visits.values())[0]} stops/day
            {customer_serving_time} visit minutes + {avg_leg_travel_min} minutes travelling time
            {sp_time_visit[salesman_ids[0]]} minutes at each stop
            Salesman can serve for total {salesman_daily_work_minutes} minutes in a day""")
        
        _DAY_NAME = {
            0 : "Monday",
            1 : "Tuesday",
            2 : "Wednesday",
            3 : "Thursday",
            4 : "Friday",
            5 : "Saturday",
            6 : "Sunday"
        }
        week_with_associated_number = sorted({date.isocalendar()[1] for date in valid_dates})
        date_to_week = {date: date.isocalendar()[1] for date in valid_dates}

        def _attempt_solve(enforce_min_visit: bool, caps_to_use: dict):
            model = cp_model.CpModel()
            
            customer_visit_as_per_segment = {
                customer_id: caps_to_use.get(customer_lookup_dict[customer_id].get("rfm_segment_final", "Low"), MIN_VISIT)
                for customer_id in customer_ids
            }
            
            visit_of_customer_by_salesman_on_that_day = {}
            for customer_id in customer_ids:
                for salesman_id in salesman_ids:
                    for date in valid_dates:
                        visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)] = model.new_bool_var(
                            f"visit_{customer_id}_{salesman_id}_{date.strftime('%Y%m%d')}"
                        )
            
            assigned_salesman_to_customer = {}
            for customer_id in customer_ids:
                for salesman_id in salesman_ids:
                    assigned_salesman_to_customer[(customer_id, salesman_id)] = model.new_bool_var(
                        f"assigned_{customer_id}_{salesman_id}"
                    )
            
            # Constraints
            # NOTE: There is NO single-salesperson assignment lock constraint here.
            # This allows multiple salespeople to be assigned to and visit the same customer
            # on different days during the month (multi-salesperson coverage enabled).
            for customer_id in customer_ids:
                for salesman_id in salesman_ids:
                    for date in valid_dates:
                        model.add(
                            visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)] <=
                            assigned_salesman_to_customer[(customer_id, salesman_id)]
                        )
            
            if truck_category.lower() == 'cold':
                cold_salesman_ids = [salesman_id for salesman_id in salesman_ids if salesman_with_cold_truck.get(salesman_id, False)]
                if cold_salesman_ids:
                    for customer_id in customer_ids:
                        for salesman_id in salesman_ids:
                            if salesman_id not in cold_salesman_ids:
                                model.add(assigned_salesman_to_customer[(customer_id, salesman_id)] == 0)
                else:
                    print("No cold capable salesman found, Allowing all SPs for cold customers as fallback.")
            
            # 3b. Every customer gets at least 1 visit
            if enforce_min_visit:
                for customer_id in customer_ids:
                    model.add(
                        sum(visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]
                            for salesman_id in salesman_ids for date in valid_dates) >= 1
                    )
                print(f"  [{territory_id}/{truck_category}] min-1-visit ENFORCED (caps = {caps_to_use})")
            else:
                print(f"  [{territory_id}/{truck_category}] min-1-visit SKIPPED/DISABLED (caps = {caps_to_use})")
            
            # 4. Max Visit per month as per the segment allows
            for customer_id in customer_ids:
                max_visit_to_customer_allowed = customer_visit_as_per_segment[customer_id]
                model.add(
                    sum(
                        visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]
                        for salesman_id in salesman_ids for date in valid_dates
                    ) <= max_visit_to_customer_allowed
                )
            
            # 5. Workload minutes constraint
            wh_overhead = max(0, wh_leg_min - avg_leg_travel_min)
            for salesman_id in salesman_ids:
                for date in valid_dates:
                    has_visits = model.new_bool_var(f"has_visit_{salesman_id}_{date.strftime('%Y%m%d')}")
                    model.add_max_equality(has_visits,
                                        [visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]
                                        for customer_id in customer_ids])
                    model.add(
                        sum((customer_serving_time + avg_leg_travel_min) * visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)] 
                            for customer_id in customer_ids) + wh_overhead * has_visits <= salesman_daily_work_minutes
                    )
            
            # 6. Each customer visited at most once per day
            for customer_id in customer_ids:
                for date in valid_dates:
                    model.add(
                        sum(visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]
                            for salesman_id in salesman_ids) <= 1
                    )
            
            # 9. Salesperson leave/holiday dates
            if group_salesman_workings_dates:
                for salesman_id in salesman_ids:
                    salesman_eligible_working_dates = set(group_salesman_workings_dates.get(salesman_id, valid_dates))
                    for date in valid_dates:
                        if date not in salesman_eligible_working_dates:
                            for customer_id in customer_ids:
                                model.add(visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)] == 0)
            
            # Soft constraints (Objective)
            objective_terms = []
            _TIER_WEIGHTS = {"High": 1_000_000_000, "Medium": 10_000_000, "Low": 100_000}
            
            for customer_id in customer_ids:
                customer_info = customer_lookup_dict[customer_id]
                customer_segment = customer_info.get("rfm_segment_final", "Low")
                customer_tier = _TIER_WEIGHTS.get(customer_segment, 100_000)
                score_bonus = int(customer_info.get('final_customer_score', 0.5) * 100)
                customer_weight = customer_tier + score_bonus
                preferred_day = customer_info.get("preferred_visit_day", "")
                
                # Proportional preferred day bonus (10% of customer weight) to strongly outweigh compactness bonus (2000)
                preferred_day_bonus = int(customer_weight * 0.1)
                
                for date in valid_dates:
                    preference_weight = preferred_day_bonus if _DAY_NAME.get(date.weekday()) == preferred_day else 0
                    for salesman_id in salesman_ids:
                        objective_terms.append(
                            (customer_weight + preference_weight) *
                            visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)])
            
            WEEKLY_CADENCE_BONUS  = 500_000_000
            ALT_WEEK_BONUS        = 5_000_000
            
            for customer_id in customer_ids:
                customer_segment = customer_lookup_dict[customer_id].get("rfm_segment_final", "Low")
                
                if customer_segment.lower() == "high":
                    for week in week_with_associated_number:
                        week_dates = [date for date in valid_dates if date_to_week[date] == week]
                        if not week_dates: continue
                        has_visit = model.new_bool_var(f"has_week_visit_{customer_id}_{week}")
                        model.add_max_equality(
                            has_visit,
                            [visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]
                            for salesman_id in salesman_ids for date in week_dates]
                        )
                        objective_terms.append(WEEKLY_CADENCE_BONUS * has_visit)
                
                elif customer_segment.lower() == "medium":
                    for week_index, week_a in enumerate(week_with_associated_number):
                        for week_b in week_with_associated_number[week_index + 1:]:
                            if week_b - week_a < 2: continue
                            dates_a = [d for d in valid_dates if date_to_week[d] == week_a]
                            dates_b = [d for d in valid_dates if date_to_week[d] == week_b]
                            if not dates_a or not dates_b: continue
                            
                            has_a = model.new_bool_var(f"mwk_{customer_id}_{week_a}")
                            has_b = model.new_bool_var(f"mwk_{customer_id}_{week_b}")
                            model.add_max_equality(
                                has_a,
                                [visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, d)] for salesman_id in salesman_ids for d in dates_a]
                            )
                            model.add_max_equality(
                                has_b,
                                [visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, d)] for salesman_id in salesman_ids for d in dates_b]
                            )
                            both_weeks = model.new_bool_var(f"mboth_{customer_id}_{week_a}_{week_b}")
                            model.add_min_equality(both_weeks, [has_a, has_b])
                            objective_terms.append(ALT_WEEK_BONUS * both_weeks)
            
            # Compactness bonus
            if customer_ids and len(salesman_ids) > 1 and len(customer_ids) >= len(salesman_ids):
                coords = np.array(
                    [
                        [
                            float(customer_lookup_dict[customer_id]["gps_lat"]),
                            float(customer_lookup_dict[customer_id]["gps_lng"])
                        ] for customer_id in customer_ids])
                
                centroids = _kmeans_centroids(coords, len(salesman_ids))
                
                salesman_to_centroid = {}
                for salesman_id in salesman_ids:
                    for k in range(len(salesman_ids)):
                        salesman_to_centroid[(salesman_id, k)] = model.new_bool_var(f"sp_cent_{salesman_id}_{k}")
                
                for salesman_id in salesman_ids:
                    model.add(sum(salesman_to_centroid[(salesman_id, k)] for k in range(len(salesman_ids))) == 1)
                for k in range(len(salesman_ids)):
                    model.add(sum(salesman_to_centroid[(sid, k)] for sid in salesman_ids) == 1)
                
                link = {}
                COMPACT_BONUS = 2000
                for customer_id in customer_ids:
                    c_lat = float(customer_lookup_dict[customer_id]["gps_lat"])
                    c_lng = float(customer_lookup_dict[customer_id]["gps_lng"])
                    proximity_c = []
                    for k, cent in enumerate(centroids):
                        dist_c = _haversine_km(c_lat, c_lng, cent[0], cent[1])
                        proximity = max(0, int((1.0 - min(dist_c / 50.0, 1.0)) * COMPACT_BONUS))
                        proximity_c.append(proximity)
                        
                    for sid in salesman_ids:
                        for k in range(len(salesman_ids)):
                            if proximity_c[k] > 0:
                                link[(customer_id, sid, k)] = model.new_bool_var(f"lnk_{customer_id}_{sid}_{k}")
                                model.add(link[(customer_id, sid, k)] <= assigned_salesman_to_customer[(customer_id, sid)])
                                model.add(link[(customer_id, sid, k)] <= salesman_to_centroid[(sid, k)])
                                objective_terms.append(proximity_c[k] * link[(customer_id, sid, k)])
            
            # Workload balancing
            if len(salesman_ids) > 1:
                sp_dates_dict = group_salesman_workings_dates if group_salesman_workings_dates else {}
                sp_total_visits = {}
                for sid in salesman_ids:
                    sp_total_visits[sid] = sum(visit_of_customer_by_salesman_on_that_day[(cid, sid, d)] for cid in customer_ids for d in valid_dates)
                
                BALANCING_PENALTY = 500
                for i, sid1 in enumerate(salesman_ids):
                    dates1 = sp_dates_dict.get(sid1, valid_dates)
                    w1 = len(dates1)
                    for sid2 in salesman_ids[i+1:]:
                        dates2 = sp_dates_dict.get(sid2, valid_dates)
                        w2 = len(dates2)
                        if w1 > 0 and w2 > 0:
                            diff = model.new_int_var(-100000, 100000, f"diff_{sid1}_{sid2}")
                            model.add(diff == sp_total_visits[sid1] * w2 - sp_total_visits[sid2] * w1)
                            abs_diff = model.new_int_var(0, 100000, f"abs_diff_{sid1}_{sid2}")
                            model.add_abs_equality(abs_diff, diff)
                            objective_terms.append(-BALANCING_PENALTY * abs_diff)
            
            model.maximize(sum(objective_terms))
            
            solver = cp_model.CpSolver()
            t_limit = solver_time if solver_time is not None else DEFAULT_SOLVER_TIME
            solver.parameters.max_time_in_seconds = t_limit
            solver.parameters.num_search_workers  = max(8, min(os.cpu_count() or 4, 16))
            solver.parameters.log_search_progress = True
            solver.parameters.log_to_stdout = False
            solver.parameters.relative_gap_limit  = 1e-4
            
            sorted_custs = sorted(customer_ids,
                                key=lambda c: customer_lookup_dict[c].get("final_customer_score", 0.5),
                                reverse=True)
            for i, cid in enumerate(sorted_custs):
                hint_sid = salesman_ids[i % len(salesman_ids)]
                for sid in salesman_ids:
                    model.add_hint(assigned_salesman_to_customer[(cid, sid)], 1 if sid == hint_sid else 0)
            
            log_path = os.path.join(os.path.dirname(__file__), "solver_log.txt")
            with open(log_path, "a", encoding="utf-8") as lf:
                lf.write(f"\n--- Solve started for Territory - {territory_id} for group - {truck_category} at {pd.Timestamp.now()} (enforce_min_1={enforce_min_visit}, caps={caps_to_use}) ---\n")
                solver.log_callback = lambda msg: lf.write(msg + "\n")
                status = solver.Solve(model)
                
            return status, solver, visit_of_customer_by_salesman_on_that_day, assigned_salesman_to_customer, customer_visit_as_per_segment

        total_slots_available = sum(sp_daily_visits[sid] * len(group_salesman_workings_dates.get(sid, valid_dates)) for sid in salesman_ids)
        can_enforce_min_visit = len(customer_ids) <= total_slots_available
        
        fallback_used = False
        caps_to_use = SEG_CAPS
        status = None
        solver = None
        visit_of_customer_by_salesman_on_that_day = {}
        assigned_salesman_to_customer = {}
        customer_visit_as_per_segment = {}
        
        if can_enforce_min_visit:
            print(f"  [{territory_id}/{truck_category}] Attempting solve with min-1-visit constraint and normal caps {SEG_CAPS}...")
            status, solver, visit_of_customer_by_salesman_on_that_day, assigned_salesman_to_customer, customer_visit_as_per_segment = _attempt_solve(
                enforce_min_visit=True,
                caps_to_use=SEG_CAPS
            )
            
            if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                print(f"  [{territory_id}/{truck_category}] Solver INFEASIBLE or FAILED with normal caps. Retrying with min-1-visit and default caps {DEFAULT_SEG_CAPS}...")
                fallback_used = True
                caps_to_use = DEFAULT_SEG_CAPS
                status, solver, visit_of_customer_by_salesman_on_that_day, assigned_salesman_to_customer, customer_visit_as_per_segment = _attempt_solve(
                    enforce_min_visit=True,
                    caps_to_use=DEFAULT_SEG_CAPS
                )
                
                if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                    print(f"  [{territory_id}/{truck_category}] Solver STILL INFEASIBLE. Retrying without min-1-visit and default caps as a last resort...")
                    status, solver, visit_of_customer_by_salesman_on_that_day, assigned_salesman_to_customer, customer_visit_as_per_segment = _attempt_solve(
                        enforce_min_visit=False,
                        caps_to_use=DEFAULT_SEG_CAPS
                    )
        else:
            print(f"  [{territory_id}/{truck_category}] min-1-visit constraint cannot be met ({len(customer_ids)} customers > {total_slots_available} slots). Solving in fallback mode directly...")
            fallback_used = True
            caps_to_use = DEFAULT_SEG_CAPS
            status, solver, visit_of_customer_by_salesman_on_that_day, assigned_salesman_to_customer, customer_visit_as_per_segment = _attempt_solve(
                enforce_min_visit=False,
                caps_to_use=DEFAULT_SEG_CAPS
            )

        if solver is not None and status is not None:
            print(f"Solving Status: {solver.status_name(status)}")
            print(f"Wall Time: {solver.wall_time:.2f} seconds")
            if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                print(f"Objective Value: {solver.objective_value}")
                print(f"Best Bound: {solver.best_objective_bound}")
            else:
                print("No feasible solution found.")
            print("\n--- Detailed Solver Statistics ---")
            print(solver.response_stats())
        else:
            print("No solver ran successfully.")
            return empty_df

        # ── Failure diagnostics ───────────────────────────────────────────────
        if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            travel_mins: dict[tuple[str, str], int] = {}
            for cid in customer_ids:
                for sid in salesman_ids: travel_mins[(cid, sid)] = avg_leg_travel_min
                
            _report_failure_diagnostics(
                solver=solver,
                status=status,
                territory_id=territory_id,
                truck_group=truck_category,
                visit=visit_of_customer_by_salesman_on_that_day,
                assigned=assigned_salesman_to_customer,
                customer_ids=customer_ids,
                sp_ids=salesman_ids,
                valid_dates=valid_dates,
                cust_lookup=customer_lookup_dict,
                sp_is_cold=salesman_with_cold_truck,
                avg_visit_minutes=customer_serving_time,
                travel_mins=travel_mins,
                daily_work_minutes=salesman_daily_work_minutes,
            )
            return empty_df

        print(f"✅ {territory_id} - {truck_category} {solver.StatusName(status)}")
        
        # Extract the solution
        rows: List[Dict[str, Any]] = []
        
        for customer_id in customer_ids:
            info = customer_lookup_dict[customer_id]
            for salesman_id in salesman_ids:
                for date in valid_dates:
                    if solver.Value(visit_of_customer_by_salesman_on_that_day[(customer_id, salesman_id, date)]) == 1:
                        wh_distance_km = _haversine_km(
                            terr_warehouse_lat,
                            terr_warehouse_lng,
                            float(info.get('gps_lat')),
                            float(info.get('gps_lng')),
                        )
                        rows.append({
                            "schedule_date":            date,
                            "sales_id":                 salesman_id,
                            "territory_id":             territory_id,
                            "customer_id":              customer_id,
                            "truck_group":              truck_category,
                            "shop_name":                info.get("shop_name", ""),
                            "locality":                 info.get("locality", ""),
                            "gps_lat":                  info.get("gps_lat"),
                            "gps_lng":                  info.get("gps_lng"),
                            "rfm_segment_final":        info.get("rfm_segment_final", "Low"),
                            "final_customer_score":     info.get("final_customer_score", 0.0),
                            "rfm_combined":             info.get("rfm_combined", 0.0),
                            "customer_rank":            info.get("customer_rank", 0),
                            "seasonality_score":        info.get("seasonality_score", 0.0),
                            "territory_score":          info.get("territory_score", 0.0),
                            "locality_score":           info.get("locality_score", 0.0),
                            "rating_score":             info.get("rating_score", 0.0),
                            "lifecycle_state":          info.get("lifecycle_state", "Active"),
                            "cold_truck_required":      info.get("cold_truck_required", False),
                            "estimated_visit_minutes":  customer_serving_time,
                            "estimated_travel_minutes": avg_leg_travel_min,
                            "warehouse_to_customer_km": round(wh_distance_km, 3),
                            "route_leg_km":             None,
                            "cumulative_route_km":      None,
                            "visit_cap":                customer_visit_as_per_segment[customer_id],
                            "fallback_used":            fallback_used,
                        })

        if not rows: return empty_df
        result_df = (
            pd.DataFrame(rows)
            .sort_values(
                ["schedule_date", "sales_id", "rfm_segment_final", "final_customer_score"],
                ascending=[True, True, True, False],
            )
            .reset_index(drop=True)
        )
        
        # ── Warn if any customer missed ───────────────────────────────────────
        scheduled = set(result_df["customer_id"].unique())
        missed    = [cid for cid in customer_ids if cid not in scheduled]
        if missed:
            print(f"""
                ⚠️Territory - {territory_id} of group - {truck_category}] has:
                {len(missed)} customer(s) NOT scheduled""")
        return result_df
    
    @staticmethod          
    def _empty_df() -> pd.DataFrame:
        return pd.DataFrame(columns=[
            "schedule_date", "sales_id", "territory_id", "customer_id", "truck_group",
            "shop_name", "locality", "gps_lat", "gps_lng",
            "rfm_segment_final", "final_customer_score", "rfm_combined",
            "customer_rank", "seasonality_score", "territory_score",
            "locality_score", "rating_score",
            "lifecycle_state", "cold_truck_required",
            "estimated_visit_minutes", "estimated_travel_minutes",
            "warehouse_to_customer_km", "route_leg_km", "cumulative_route_km",
            "visit_cap", "fallback_used",
        ])
    
class DailyRoutePlanner:
    """
        Given a day's list of customers (already assigned to a salesperson),
        orders them using Google OR-Tools Routing Solver (VRP/TSP), then computes:
            route_leg_km        — one-way distance from previous stop
            cumulative_route_km — running total km for the day
    """
    def get_route(
        self,
        day_schedule: pd.DataFrame,
        start_lat: float,
        start_lng: float,
    ) -> pd.DataFrame:
        if day_schedule.empty:
            return day_schedule.copy()

        df_c = day_schedule.copy().reset_index(drop=True)
        n_cust = len(df_c)

        if n_cust == 1:
            row = df_c.loc[0].to_dict()
            leg_km = _haversine_km(start_lat, start_lng, float(row["gps_lat"]), float(row["gps_lng"]))
            row["route_leg_km"] = round(leg_km, 3)
            row["cumulative_route_km"] = round(leg_km, 3)
            result = pd.DataFrame([row])
            result["route_rank"] = [1]
            return result

        # Setup node coordinates mapping: node 0 = warehouse, node 1..N = customers
        def _get_node_coords(node_idx: int) -> tuple[float, float]:
            if node_idx <= 0 or node_idx > n_cust:
                return (start_lat, start_lng)
            row_c = df_c.loc[node_idx - 1]
            return (float(row_c["gps_lat"]), float(row_c["gps_lng"]))

        # Manager and Routing Model
        n_nodes = n_cust + 1
        manager = pywrapcp.RoutingIndexManager(n_nodes, 1, 0)
        routing = pywrapcp.RoutingModel(manager)

        # Distance callback: distance from any node to node 0 (warehouse) is 0 to enforce one-way
        def distance_callback(from_index, to_index):
            try:
                from_node = manager.IndexToNode(from_index)
                to_node = manager.IndexToNode(to_index)
                if to_node <= 0 or to_node > n_cust:
                    return 0
                lat_i, lng_i = _get_node_coords(from_node)
                lat_j, lng_j = _get_node_coords(to_node)
                # multiply by 1000 to get meters for integer precision
                return int(_haversine_km(lat_i, lng_i, lat_j, lng_j) * 1000)
            except Exception as e:
                print(f"  [Warning] Exception in distance callback for index ({from_index} -> {to_index}): {e}")
                return 999999

        transit_callback_index = routing.RegisterTransitCallback(distance_callback)
        routing.SetArcCostEvaluatorOfAllVehicles(transit_callback_index)

        # Solving Parameters
        search_parameters = pywrapcp.DefaultRoutingSearchParameters()
        search_parameters.first_solution_strategy = (
            routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
        )


        """
            Commented out the Guided Local Path
        """
        
        '''
        search_parameters.local_search_metaheuristic = (
            routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
        )
        search_parameters.time_limit.seconds = 1  # solve very
        '''

        # Note: We do NOT set local_search_metaheuristic to GUIDED_LOCAL_SEARCH.
        # Leaving it default (UNSET) allows the solver to run standard local search
        # and terminate immediately once it reaches a local minimum (usually in milliseconds),
        # instead of wasting the entire time limit trying to escape local minima.
        # This keeps the overall schedule generation execution time extremely fast.
        search_parameters.time_limit.seconds = 2  # Safety guard, terminates early in milliseconds

        solution = routing.SolveWithParameters(search_parameters)

        if not solution:
            # Fallback to simple nearest neighbour if routing solver fails
            status_code = routing.status()
            status_name = {
                0: "ROUTING_NOT_SOLVED",
                1: "ROUTING_SUCCESS",
                2: "ROUTING_FAIL",
                3: "ROUTING_FAIL_TIMEOUT",
                4: "ROUTING_INVALID"
            }.get(status_code, f"UNKNOWN ({status_code})")
            print(f"  ⚠️  Routing solver failed to find solution (Status: {status_name}). Falling back to Nearest Neighbour.")
            return self._nn_fallback(day_schedule, start_lat, start_lng)

        # Reconstruct route from solution
        index = routing.Start(0)
        route_indices = []
        while not routing.IsEnd(index):
            node = manager.IndexToNode(index)
            if node > 0:
                route_indices.append(node - 1)
            index = solution.Value(routing.NextVar(index))

        # Reorder DataFrame rows based on optimal VRP route
        ordered_df = df_c.iloc[route_indices].copy().reset_index(drop=True)

        # Compute legs and cumulative km
        route: list[dict] = []
        cur_lat, cur_lng = start_lat, start_lng
        cumulative_km = 0.0

        for idx, row in ordered_df.iterrows():
            leg_km = _haversine_km(cur_lat, cur_lng, float(row["gps_lat"]), float(row["gps_lng"]))
            cumulative_km += leg_km
            row_dict = row.to_dict()
            row_dict["route_leg_km"] = round(leg_km, 3)
            row_dict["cumulative_route_km"] = round(cumulative_km, 3)
            route.append(row_dict)
            cur_lat, cur_lng = float(row["gps_lat"]), float(row["gps_lng"])

        result = pd.DataFrame(route)
        result["route_rank"] = range(1, len(result) + 1)
        return result

    def _nn_fallback(
        self,
        day_schedule: pd.DataFrame,
        start_lat: float,
        start_lng: float,
    ) -> pd.DataFrame:
        unvisited = day_schedule.copy().reset_index(drop=True)
        route: list[dict] = []
        cur_lat, cur_lng = start_lat, start_lng
        cumulative_km = 0.0

        while not unvisited.empty:
            distances = unvisited.apply(
                lambda r: _haversine_km(cur_lat, cur_lng, float(r["gps_lat"]), float(r["gps_lng"])),
                axis=1,
            )
            nearest_idx = int(distances.idxmin())
            row         = unvisited.loc[nearest_idx].to_dict()
            leg_km      = distances[nearest_idx]
            cumulative_km += leg_km
            row["route_leg_km"]        = round(leg_km, 3)
            row["cumulative_route_km"] = round(cumulative_km, 3)
            route.append(row)
            cur_lat = float(row["gps_lat"])
            cur_lng = float(row["gps_lng"])
            unvisited = unvisited.drop(index=nearest_idx).reset_index(drop=True)

        result = pd.DataFrame(route)
        result["route_rank"] = range(1, len(result) + 1)
        return result
    
# ─────────────────────────────────────────────────────────────────────────────
# Folium map helpers  (updated for truck_group, km, new fields)
# ─────────────────────────────────────────────────────────────────────────────

_SP_COLOURS = [
    "red", "blue", "green", "purple", "orange",
    "darkred", "cadetblue", "darkgreen", "darkpurple", "black",
]
SEGMENT_COLOUR = {"High": "red", "Medium": "orange", "Low": "gray"}


def _add_warehouse_marker(m, wh_lat: float, wh_lng: float, label: str = "Warehouse"):
    import folium
    folium.Marker(
        [wh_lat, wh_lng],
        icon=folium.Icon(color="black", icon="home", prefix="fa"),
        tooltip=f"🏭 {label} (start of route)",
        popup=f"<b>{label}</b><br>Route start",
    ).add_to(m)


def build_route_map_for_salesperson(
    daily_schedule:    pd.DataFrame,
    detailed_schedule: pd.DataFrame,
    sales_id:          str,
    schedule_date:     str | pd.Timestamp,
    warehouse_lat:     float = 0.0,
    warehouse_lng:     float = 0.0,
    zoom_start:        int   = 12,
):
    """
    Folium map for ONE salesperson on ONE day.
    Tooltip shows RFM segment, score, rank, and route leg km.
    """
    try:
        import folium
        from folium.plugins import PolyLineTextPath
    except ImportError:
        raise ImportError("pip install folium")

    sched_date = pd.Timestamp(schedule_date).normalize()
    day_df = detailed_schedule[
        (detailed_schedule["sales_id"]      == sales_id) &
        (detailed_schedule["schedule_date"] == sched_date)
    ].copy()

    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank")
    day_df = day_df.reset_index(drop=True)

    if day_df.empty:
        raise ValueError(f"No visits for {sales_id} on {sched_date.date()}")

    center_lat = day_df["gps_lat"].mean()
    center_lng = day_df["gps_lng"].mean()
    m = folium.Map(location=[center_lat, center_lng], zoom_start=zoom_start)

    if warehouse_lat != 0.0 or warehouse_lng != 0.0:
        _add_warehouse_marker(m, warehouse_lat, warehouse_lng)
        route_coords = [[warehouse_lat, warehouse_lng]] + day_df[["gps_lat", "gps_lng"]].values.tolist()
    else:
        route_coords = day_df[["gps_lat", "gps_lng"]].values.tolist()

    for _, row in day_df.iterrows():
        color  = SEGMENT_COLOUR.get(str(row.get("rfm_segment_final", "")), "blue")
        score  = row.get("final_customer_score", 0.0)
        rank   = row.get("customer_rank", "?")
        leg_km = row.get("route_leg_km", 0.0)
        folium.Marker(
            [row["gps_lat"], row["gps_lng"]],
            icon=folium.DivIcon(
                html=f"""<div style="background:{color};color:white;border-radius:50%;
                         width:26px;height:26px;text-align:center;line-height:26px;
                         font-weight:bold;font-size:11px;">{row.get('route_rank','')}</div>""",
            ),
            tooltip=(
                f"{row.get('route_rank','')}. {row.get('shop_name','')} "
                f"[{row.get('rfm_segment_final','')}] "
                f"score={score:.3f} rank=#{rank} "
                f"leg={leg_km:.2f}km "
                f"cold={row.get('cold_truck_required','?')}"
            ),
        ).add_to(m)

    line = folium.PolyLine(route_coords, weight=3, color="navy")
    m.add_child(line)
    PolyLineTextPath(line, "➤", repeat=True, offset=7,
                     attributes={"fill": "red", "font-size": "14"}).add_to(m)

    m.get_root().html.add_child(folium.Element(f"""
    <div style="position:fixed;top:10px;left:60px;z-index:1000;
                background:white;padding:8px 12px;border-radius:6px;
                border:1px solid #ccc;font-family:sans-serif;">
        <b>{sales_id}</b> — {sched_date.strftime('%d %b %Y')} &nbsp;|&nbsp; {len(day_df)} stops
    </div>"""))

    return m


def build_territory_day_map(
    result:        MultiScheduleResult,
    territory_id:  str,
    schedule_date: str | pd.Timestamp,
    truck_group:   Optional[str] = None,   # None → show both cold + normal
    zoom_start:    int = 11,
):
    """
    Shows all salespeople in one territory on a single folium map for a given date.
    truck_group filter: "cold", "normal", or None (both).
    """
    try:
        import folium
        from folium.plugins import PolyLineTextPath
    except ImportError:
        raise ImportError("pip install folium")

    sched_date = pd.Timestamp(schedule_date).normalize()

    # Filter detailed schedule
    df = result.detailed_schedule.copy()
    df = df[
        (df["territory_id"] == territory_id) &
        (df["schedule_date"] == sched_date)
    ]
    if truck_group:
        df = df[df["truck_group"] == truck_group]

    if df.empty:
        raise ValueError(f"No visits in territory {territory_id} on {sched_date.date()} "
                         f"(truck_group={truck_group})")

    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))

    center_lat = df["gps_lat"].mean()
    center_lng = df["gps_lng"].mean()
    m = folium.Map(location=[center_lat, center_lng], zoom_start=zoom_start)

    if wh_lat != 0.0 or wh_lng != 0.0:
        _add_warehouse_marker(m, wh_lat, wh_lng, label=f"Warehouse ({territory_id})")

    sp_ids = sorted(df["sales_id"].unique())
    sp_colour_map = {sid: _SP_COLOURS[i % len(_SP_COLOURS)] for i, sid in enumerate(sp_ids)}
    legend_items  = []

    for sid in sp_ids:
        sp_df     = df[df["sales_id"] == sid].copy()
        if "route_rank" in sp_df.columns:
            sp_df = sp_df.sort_values("route_rank")
        sp_color  = sp_colour_map[sid]

        route_coords = (
            [[wh_lat, wh_lng]] + sp_df[["gps_lat", "gps_lng"]].values.tolist()
            if (wh_lat != 0.0 or wh_lng != 0.0)
            else sp_df[["gps_lat", "gps_lng"]].values.tolist()
        )

        for _, row in sp_df.iterrows():
            score  = row.get("final_customer_score", 0.0)
            leg_km = row.get("route_leg_km", 0.0)
            folium.Marker(
                [row["gps_lat"], row["gps_lng"]],
                icon=folium.DivIcon(
                    html=f"""<div style="background:{sp_color};color:white;
                             border-radius:50%;width:26px;height:26px;
                             text-align:center;line-height:26px;
                             font-weight:bold;font-size:11px;
                             border:2px solid white;">{row.get('route_rank','')}</div>""",
                ),
                tooltip=(
                    f"[{sid}] {row.get('route_rank','')}. {row.get('shop_name','')} "
                    f"[{row.get('rfm_segment_final','')}] "
                    f"score={score:.3f} leg={leg_km:.2f}km "
                    f"cold={row.get('cold_truck_required','?')}"
                ),
            ).add_to(m)

        line = folium.PolyLine(route_coords, weight=3, color=sp_color, opacity=0.8)
        m.add_child(line)
        PolyLineTextPath(line, "➤", repeat=True, offset=7,
                         attributes={"fill": sp_color, "font-size": "13"}).add_to(m)

        seg_str = " | ".join(
            f"{k}:{v}" for k, v in sp_df["rfm_segment_final"].value_counts().items()
        )
        total_km = sp_df["route_leg_km"].sum() if "route_leg_km" in sp_df.columns else 0
        legend_items.append(
            f'<li><span style="background:{sp_color};width:14px;height:14px;'
            f'display:inline-block;border-radius:50%;margin-right:6px;"></span>'
            f'{sid} ({len(sp_df)} stops, {total_km:.1f} km — {seg_str})</li>'
        )

    legend_html = f"""
    <div style="position:fixed;top:10px;right:10px;z-index:1000;
                background:white;padding:10px 14px;border-radius:8px;
                border:1px solid #ccc;font-family:sans-serif;font-size:12px;
                min-width:240px;">
        <b>{territory_id}</b> — {sched_date.strftime('%d %b %Y')}
        {(' | 🧊 Cold only' if truck_group=='cold' else ' | 🚚 Normal only' if truck_group=='normal' else '')}<br>
        <ul style="list-style:none;margin:6px 0 0;padding:0;">
            {''.join(legend_items)}
            <li style="margin-top:6px;">
                <span style="background:black;width:14px;height:14px;
                display:inline-block;border-radius:50%;margin-right:6px;"></span>
                Warehouse
            </li>
        </ul>
        <hr style="margin:6px 0;">
        <span style="font-size:11px;color:#666;">
            Marker colour = salesperson &nbsp;|&nbsp; Number = route stop<br>
            High=red &nbsp; Medium=orange &nbsp; Low=gray (segment in tooltip)
        </span>
    </div>"""
    m.get_root().html.add_child(folium.Element(legend_html))
    return m


# ─────────────────────────────────────────────────────────────────────────────
# Stop-to-stop distance: Folium map + Excel export
# ─────────────────────────────────────────────────────────────────────────────

def build_stop_to_stop_distance_table(
    result,
    territory_id:  str,
    schedule_date,
    sales_id:      str,
) -> pd.DataFrame:
    """
    Returns a DataFrame with one row per consecutive stop pair on the route:
        stop_from | stop_to | from_id | from_shop | to_id | to_shop | leg_km | leg_min
    Row 0 is always Warehouse → Stop 1.
    """
    sched_date = pd.Timestamp(schedule_date).normalize()
    det = result.detailed_schedule
    day_df = det[
        (det["territory_id"]  == territory_id) &
        (det["schedule_date"] == sched_date) &
        (det["sales_id"]      == sales_id)
    ].copy()

    if day_df.empty:
        return pd.DataFrame()

    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank").reset_index(drop=True)

    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))
    avg_speed = 32.0   # km/h

    stops = [{"stop_num": 0, "customer_id": "WAREHOUSE", "shop_name": "Warehouse",
               "gps_lat": wh_lat, "gps_lng": wh_lng}]
    for _, row in day_df.iterrows():
        stops.append({"stop_num": int(row.get("route_rank", len(stops))),
                      "customer_id": row["customer_id"],
                      "shop_name": row.get("shop_name", ""),
                      "gps_lat": float(row["gps_lat"]), "gps_lng": float(row["gps_lng"])})

    rows = []
    for i in range(len(stops) - 1):
        f, t = stops[i], stops[i + 1]
        km = _haversine_km(f["gps_lat"], f["gps_lng"], t["gps_lat"], t["gps_lng"])
        rows.append({
            "stop_from":  f["stop_num"],  "stop_to":   t["stop_num"],
            "from_id":    f["customer_id"], "from_shop": f["shop_name"],
            "from_lat":   round(f["gps_lat"], 6), "from_lng": round(f["gps_lng"], 6),
            "to_id":      t["customer_id"], "to_shop":   t["shop_name"],
            "to_lat":     round(t["gps_lat"], 6), "to_lng": round(t["gps_lng"], 6),
            "leg_km":     round(km, 3),
            "leg_min":    round((km / avg_speed) * 60, 1),
        })
    return pd.DataFrame(rows)


def build_stop_to_stop_map(
    result,
    territory_id:  str,
    schedule_date,
    sales_id:      str,
    zoom_start:    int = 13,
):
    """
    Folium map with stop-to-stop legs labelled with distance (km).
    Warehouse is the start; each polyline leg shows km at its midpoint.
    """
    try:
        import folium
    except ImportError:
        raise ImportError("pip install folium")

    dist_df    = build_stop_to_stop_distance_table(result, territory_id, schedule_date, sales_id)
    sched_date = pd.Timestamp(schedule_date).normalize()
    if dist_df.empty:
        raise ValueError(f"No route for {sales_id} in {territory_id} on {schedule_date}")

    all_lats = dist_df["from_lat"].tolist() + [dist_df["to_lat"].iloc[-1]]
    all_lngs = dist_df["from_lng"].tolist() + [dist_df["to_lng"].iloc[-1]]
    m = folium.Map(location=[sum(all_lats)/len(all_lats), sum(all_lngs)/len(all_lngs)],
                   zoom_start=zoom_start)

    # Draw each leg + distance label at midpoint
    for _, leg in dist_df.iterrows():
        coords = [[leg["from_lat"], leg["from_lng"]], [leg["to_lat"], leg["to_lng"]]]
        folium.PolyLine(coords, weight=3, color="navy", opacity=0.85).add_to(m)
        mid_lat = (leg["from_lat"] + leg["to_lat"]) / 2
        mid_lng = (leg["from_lng"] + leg["to_lng"]) / 2
        folium.Marker(
            [mid_lat, mid_lng],
            icon=folium.DivIcon(
                html=(
                    '<div style="background:rgba(255,255,255,0.88);padding:1px 5px;'
                    'border-radius:3px;font-size:10px;color:#003;white-space:nowrap;">'
                    + f'{leg["leg_km"]:.2f} km</div>'
                ),
                icon_size=(70, 18), icon_anchor=(35, 9),
            ),
            tooltip=f'Leg {int(leg["stop_from"])}\u2192{int(leg["stop_to"])}: {leg["leg_km"]:.2f} km ({leg["leg_min"]:.0f} min)',
        ).add_to(m)

    # Stop markers
    wh_lat, wh_lng = result.territory_warehouses.get(territory_id, (0.0, 0.0))
    _add_warehouse_marker(m, wh_lat, wh_lng)

    det    = result.detailed_schedule
    day_df = det[
        (det["territory_id"]  == territory_id) &
        (det["schedule_date"] == sched_date) &
        (det["sales_id"]      == sales_id)
    ].copy()
    if "route_rank" in day_df.columns:
        day_df = day_df.sort_values("route_rank")

    for _, row in day_df.iterrows():
        color = SEGMENT_COLOUR.get(str(row.get("rfm_segment_final", "")), "blue")
        folium.Marker(
            [row["gps_lat"], row["gps_lng"]],
            icon=folium.DivIcon(
                html=(
                    f'<div style="background:{color};color:white;border-radius:50%;'
                    f'width:26px;height:26px;text-align:center;line-height:26px;'
                    f'font-weight:bold;font-size:11px;">{row.get("route_rank","")}</div>'
                ),
            ),
            tooltip=(
                f'{row.get("route_rank","")}. {row.get("shop_name","")} '
                f'[{row.get("rfm_segment_final","")}] '
                f'leg={row.get("route_leg_km", 0):.2f} km'
            ),
        ).add_to(m)

    total_km = dist_df["leg_km"].sum()
    m.get_root().html.add_child(folium.Element(
        f'<div style="position:fixed;top:10px;left:60px;z-index:1000;'
        f'background:white;padding:8px 12px;border-radius:6px;'
        f'border:1px solid #ccc;font-family:sans-serif;">'
        f'<b>{sales_id}</b> \u2014 {sched_date.strftime("%d %b %Y")}'
        f'&nbsp;|&nbsp;{len(dist_df)} legs&nbsp;|&nbsp;Total: {total_km:.2f} km'
        f'</div>'
    ))
    return m


def export_stop_to_stop_excel(
    result,
    territory_id:  str,
    schedule_date,
    filepath:      str = "D:/Data Science/Basamh/JP_Yash/journey-planner/jp_data_output/stop_to_stop_distances.xlsx",
):
    """
    Excel workbook — one sheet per salesperson, one row per route leg.
    Columns: stop_from | stop_to | from_shop | to_shop | leg_km | leg_min
    Sheet 'Summary' gives total km per salesperson.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    sched_date = pd.Timestamp(schedule_date).normalize()
    det = result.detailed_schedule
    sp_list = sorted(
        det[(det["territory_id"]  == territory_id) &
            (det["schedule_date"] == sched_date)]["sales_id"].unique()
    )
    if not sp_list:
        raise ValueError(f"No schedule for {territory_id} on {sched_date.date()}")

    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT   = Font(color="FFFFFF", bold=True)
    ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")
    TOTAL_FONT = Font(bold=True)

    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = []

    for sid in sp_list:
        dist_df = build_stop_to_stop_distance_table(result, territory_id, sched_date, sid)
        if dist_df.empty:
            continue

        ws = wb.create_sheet(title=sid[-12:])
        headers = ["From Stop", "To Stop", "From ID", "From Shop",
                   "From Lat", "From Lng", "To ID", "To Shop",
                   "To Lat", "To Lng", "Leg km", "Leg min"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

        for ri, row in dist_df.iterrows():
            er   = ri + 2
            fill = ALT_FILL if ri % 2 == 0 else None
            vals = [int(row["stop_from"]), int(row["stop_to"]),
                    row["from_id"], row["from_shop"],
                    row["from_lat"], row["from_lng"],
                    row["to_id"],   row["to_shop"],
                    row["to_lat"],  row["to_lng"],
                    round(float(row["leg_km"]), 3),
                    round(float(row["leg_min"]), 1)]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=er, column=ci, value=val)
                if fill:
                    cell.fill = fill

        tr = len(dist_df) + 2
        for ci, val in [(10, "TOTAL"),
                        (11, round(float(dist_df["leg_km"].sum()), 3)),
                        (12, round(float(dist_df["leg_min"].sum()), 1))]:
            ws.cell(row=tr, column=ci, value=val).font = TOTAL_FONT

        col_widths = [10, 8, 16, 30, 10, 10, 16, 30, 10, 10, 10, 10]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        summary_rows.append({
            "sales_id":  sid,
            "n_stops":   len(dist_df),
            "total_km":  round(float(dist_df["leg_km"].sum()), 3),
            "total_min": round(float(dist_df["leg_min"].sum()), 1),
        })

    # Summary sheet
    ws_s = wb.create_sheet(title="Summary", index=0)
    for ci, h in enumerate(["Salesperson", "Stops", "Total km", "Total min"], 1):
        c = ws_s.cell(row=1, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(summary_rows, 2):
        ws_s.cell(row=ri, column=1, value=row["sales_id"])
        ws_s.cell(row=ri, column=2, value=row["n_stops"])
        ws_s.cell(row=ri, column=3, value=row["total_km"])
        ws_s.cell(row=ri, column=4, value=row["total_min"])
    for ci, w in enumerate([22, 10, 12, 12], 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.cell(row=1, column=6,
              value=f"{territory_id}  |  {sched_date.strftime('%d %b %Y')}").font = Font(bold=True)

    wb.save(filepath)
    print(f"Saved stop-to-stop Excel \u2192 {filepath}  ({len(summary_rows)} salesperson sheet(s))")
    return filepath

def export_under_visited_excel(
    result,
    filepath: str = r"D:/Data Science/Basamh/JP_Yash/journey-planner/jp_data_output/under_visited_customers.xlsx",
    territory_id: str = None,
) -> str:
    """
    Exports a 4-sheet Excel workbook that shows exactly how well the schedule
    served each customer relative to their segment visit cap.

    Sheets
    ──────
    Summary                — KPI cards, per-segment table, capacity note
    Under-Visited Customers — every customer with >0 but <cap visits
    Full Report            — all customers: actual vs cap, colour-coded
    Segment x SP           — delivery breakdown by salesperson × segment

    Extra columns added to Under-Visited / Full Report
    ───────────────────────────────────────────────────
    actual_visits   — visits the customer received this month
    visit_cap       — segment cap SEG_CAPS
    visits_gap      — cap − actual  (how many visits were missed)
    pct_of_cap      — actual / cap × 100

    Usage
    ─────
    from scheduler_4 import export_under_visited_excel

    export_under_visited_excel(result, "under_visited.xlsx", territory_id="TER_RUH")
    # or for all territories:
    export_under_visited_excel(result, "under_visited.xlsx")

    Returns the filepath that was written.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        raise ImportError("pip install openpyxl")

    combined = result.detailed_schedule
    if combined.empty:
        print("No schedule data to export.")
        return filepath

    _CAPS = SEG_CAPS

    # ── Build per-customer base table ────────────────────────────────────────
    visit_counts = (
        combined.groupby("customer_id")
        .size()
        .rename("actual_visits")
        .reset_index()
    )
    keep_cols = [c for c in
                 ["customer_id", "rfm_segment_final", "final_customer_score",
                  "sales_id", "territory_id", "cold_truck_required",
                  "lifecycle_state", "shop_name"]
                 if c in combined.columns]
    base = (
        combined[keep_cols]
        .drop_duplicates("customer_id")
        .merge(visit_counts, on="customer_id", how="left")
    )
    base["actual_visits"] = base["actual_visits"].fillna(0).astype(int)
    base["visit_cap"]     = base["rfm_segment_final"].map(_CAPS).fillna(1).astype(int)
    base["visits_gap"]    = base["visit_cap"] - base["actual_visits"]
    base["pct_of_cap"]    = (base["actual_visits"] / base["visit_cap"] * 100).round(1)

    if territory_id:
        base = base[base["territory_id"] == territory_id].copy()

    under_df = (
        base[(base["actual_visits"] > 0) & (base["visits_gap"] > 0)]
        .sort_values(["rfm_segment_final", "visits_gap"], ascending=[True, False])
        .reset_index(drop=True)
    )
    full_df = base.sort_values(
        ["rfm_segment_final", "visits_gap"], ascending=[True, False]
    ).reset_index(drop=True)

    seg_summary = full_df.groupby("rfm_segment_final").agg(
        customers=("customer_id", "count"),
        actual_visits=("actual_visits", "sum"),
        max_possible=("visit_cap", "sum"),
        avg_actual=("actual_visits", "mean"),
        under_visited=("visits_gap", lambda x: (x > 0).sum()),
        total_gap=("visits_gap", "sum"),
        avg_pct=("pct_of_cap", "mean"),
    ).round(2).reset_index()

    working_days = combined["schedule_date"].nunique()
    total_slots  = int(combined.groupby(["sales_id", "schedule_date"]).ngroups
                       * combined.groupby("sales_id")
                         .apply(lambda g: g.groupby("schedule_date").size().mean()).mean())
    # simpler: actual total visits
    total_actual = len(combined)
    total_demand = int(full_df["visit_cap"].sum())
    scope_label  = f"Territory: {territory_id}" if territory_id else "All Territories"

    # ── Style helpers ────────────────────────────────────────────────────────
    thin   = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    SEG_FILL = {
        "High":   PatternFill("solid", fgColor="FCE4D6"),
        "Medium": PatternFill("solid", fgColor="FFF2CC"),
        "Low":    PatternFill("solid", fgColor="F2F2F2"),
    }
    SEG_FG  = {"High": "C00000", "Medium": "7F4F24", "Low": "404040"}
    ALT     = PatternFill("solid", fgColor="F9F9F9")

    def _hdr(ws, row, values, bg, fg="FFFFFF"):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(color=fg, bold=True, size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[row].height = 22

    def _cell(ws, r, c, v, fill=None, bold=False, align="left", fg="000000"):
        cell = ws.cell(row=r, column=c, value=v)
        if fill:         cell.fill = fill
        cell.font      = Font(bold=bold, size=10, name="Calibri", color=fg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = BORDER
        ws.row_dimensions[r].height = 16
        return cell

    def _widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _freeze(ws, ref="A3"):
        ws.freeze_panes    = ref
        ws.auto_filter.ref = ws.dimensions

    ROW_COLS = ["Customer ID", "Shop Name", "Segment", "Assigned SP", "Territory",
                "Cold Truck", "Lifecycle", "Actual Visits", "Visit Cap",
                "Visits Gap", "% of Cap"]

    def _write_customer_rows(ws, df, start_row=3):
        for ri, (_, row) in enumerate(df.iterrows()):
            r   = ri + start_row
            seg = row["rfm_segment_final"]
            vals = [
                row["customer_id"],
                row.get("shop_name", ""),
                seg,
                row.get("sales_id", ""),
                row.get("territory_id", ""),
                "Yes" if row.get("cold_truck_required") else "No",
                row.get("lifecycle_state", ""),
                int(row["actual_visits"]),
                int(row["visit_cap"]),
                int(row["visits_gap"]),
                f"{row['pct_of_cap']:.1f}%",
            ]
            gap_fill = (
                PatternFill("solid", fgColor="FCE4D6") if row["visits_gap"] >= 20
                else PatternFill("solid", fgColor="FFF2CC") if row["visits_gap"] >= 8
                else ALT
            )
            for c, v in enumerate(vals, 1):
                fill = (
                    gap_fill if c == 10
                    else SEG_FILL.get(seg) if c == 3
                    else (ALT if ri % 2 == 0 else None)
                )
                _cell(ws, r, c, v, fill=fill, bold=(c == 1),
                      align="center" if c > 2 else "left",
                      fg=SEG_FG.get(seg, "000000") if c == 3 else "000000")
        if len(df) > 0:
            ws.conditional_formatting.add(
                f"K{start_row}:K{start_row - 1 + len(df)}",
                ColorScaleRule(
                    start_type="num", start_value=0,   start_color="F8696B",
                    mid_type="num",   mid_value=50,    mid_color="FFEB84",
                    end_type="num",   end_value=100,   end_color="63BE7B",
                )
            )

    # ═══ Build workbook ═══════════════════════════════════════════════════════
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:H1")
    t1 = ws1["A1"]
    t1.value     = f"Monthly Schedule — Visit Delivery Report  ({scope_label})"
    t1.font      = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    t1.alignment = Alignment(horizontal="center", vertical="center")
    t1.fill      = PatternFill("solid", fgColor="DEEAF1")
    ws1.row_dimensions[1].height = 30

    # KPI row
    kpis = [
        ("Total Customers",      len(full_df),   "1F4E79"),
        ("Working Days",         working_days,   "1F4E79"),
        ("Actual Visits",        total_actual,   "2E7D32"),
        ("Demand (full caps)",   total_demand,   "C00000"),
        ("Unvisited",            len(result.unvisited_customers) if hasattr(result, "unvisited_customers") else 0, "C00000"),
        ("Under-Visited",        len(under_df),  "E65100"),
    ]
    for i, (label, val, clr) in enumerate(kpis):
        col = i + 1
        cl  = get_column_letter(col)
        ws1.merge_cells(f"{cl}3:{cl}4")
        ws1.merge_cells(f"{cl}5:{cl}6")
        lc = ws1[f"{cl}3"]
        lc.value     = label
        lc.font      = Font(size=9, color="666666", name="Calibri")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws1[f"{cl}5"]
        vc.value     = val
        vc.font      = Font(bold=True, size=18, color=clr, name="Calibri")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    # Segment table
    r = 8
    _hdr(ws1, r,
         ["Segment", "Customers", "Cap / Customer", "Max Possible",
          "Actual Visits", "Avg % of Cap", "Under-Visited", "Total Missed Visits"],
         "1F4E79")
    for _, row in seg_summary.iterrows():
        r += 1
        seg = row["rfm_segment_final"]
        for c, v in enumerate(
            [seg, int(row["customers"]), _CAPS.get(seg, 1),
             int(row["max_possible"]), int(row["actual_visits"]),
             f"{row['avg_pct']:.1f}%", int(row["under_visited"]),
             int(row["total_gap"])], 1
        ):
            _cell(ws1, r, c, v,
                  fill=SEG_FILL.get(seg, ALT),
                  bold=(c == 1),
                  align="center" if c > 1 else "left",
                  fg=SEG_FG.get(seg, "000000") if c == 1 else "000000")

    # Capacity note
    r += 2
    ws1.merge_cells(f"A{r}:H{r}")
    note = ws1[f"A{r}"]
    note.value = (
        f"Note: min-1-visit is NOT a hard constraint. "
        f"Demand ({total_demand}) >> Capacity ({total_actual}). "
        f"Solver fills slots by priority weight. "
        f"Unvisited and under-visited customers reflect real capacity overflow — not solver errors."
    )
    note.font      = Font(italic=True, size=10, color="C00000", name="Calibri")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border    = BORDER
    ws1.row_dimensions[r].height = 28
    _widths(ws1, [18, 12, 14, 16, 14, 14, 14, 20])

    # ── Sheet 2: Under-Visited ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Under-Visited Customers")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:K1")
    t2 = ws2["A1"]
    t2.value     = f"Under-Visited Customers — Received > 0 but < Cap  ({scope_label})"
    t2.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    t2.fill      = PatternFill("solid", fgColor="C00000")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26
    _hdr(ws2, 2, ROW_COLS, "404040")
    _write_customer_rows(ws2, under_df, start_row=3)
    _freeze(ws2)
    _widths(ws2, [16, 38, 10, 14, 12, 10, 12, 12, 10, 10, 10])

    # ── Sheet 3: Full Report ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Full Report")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:K1")
    t3 = ws3["A1"]
    t3.value     = f"Full Customer Visit Report — All Customers  ({scope_label})"
    t3.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    t3.fill      = PatternFill("solid", fgColor="1F4E79")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26
    _hdr(ws3, 2, ROW_COLS, "1F4E79")
    _write_customer_rows(ws3, full_df, start_row=3)
    _freeze(ws3)
    _widths(ws3, [16, 38, 10, 14, 12, 10, 12, 12, 10, 10, 10])

    # ── Sheet 4: Segment × SP ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Segment x SP")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value     = f"Visit Delivery by Salesperson × Segment  ({scope_label})"
    t4.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    t4.fill      = PatternFill("solid", fgColor="1F4E79")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 26

    sp_seg = (
        full_df.groupby(["sales_id", "rfm_segment_final"])
        .agg(customers=("customer_id", "count"),
             actual=("actual_visits", "sum"),
             cap_total=("visit_cap", "sum"),
             avg_visits=("actual_visits", "mean"),
             avg_pct=("pct_of_cap", "mean"))
        .round(2).reset_index()
    )
    _hdr(ws4, 2,
         ["Salesperson", "Segment", "Customers", "Actual Visits",
          "Max Possible", "Avg Visits", "Avg % of Cap"],
         "1F4E79")
    for ri, (_, row) in enumerate(sp_seg.iterrows()):
        r   = ri + 3
        seg = row["rfm_segment_final"]
        for c, v in enumerate(
            [row["sales_id"], seg, int(row["customers"]), int(row["actual"]),
             int(row["cap_total"]), f"{row['avg_visits']:.1f}", f"{row['avg_pct']:.1f}%"], 1
        ):
            _cell(ws4, r, c, v,
                  fill=SEG_FILL.get(seg, ALT) if c == 2 else (ALT if ri % 2 == 0 else None),
                  bold=(c == 1),
                  align="center" if c > 1 else "left",
                  fg=SEG_FG.get(seg, "000000") if c == 2 else "000000")
    _freeze(ws4)
    _widths(ws4, [16, 12, 12, 14, 14, 12, 14])

    wb.save(filepath)
    print(f"Saved → {filepath}  "
          f"({len(under_df)} under-visited | {len(full_df)} total customers)")
    return filepath